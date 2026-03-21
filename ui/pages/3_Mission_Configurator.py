"""
ui/pages/3_Mission_Configurator.py
Select batteries, mission profile, and UAV configuration.
Also provides a Mission Builder to create and edit missions and phases.
"""
import sys
from pathlib import Path
_PROJECT_ROOT = Path(__file__).parent.parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt

st.set_page_config(page_title="Mission Configurator", page_icon="⚙️", layout="wide")

from ui.components.db_helpers import load_db, load_config, save_config, plot_phase_power, reload_db
from ui.config import PHASE_COLORS, DB_PATH

st.title("⚙️ Mission Configurator")
st.caption("Configure your analysis setup, or build and edit mission profiles.")

db = load_db()

tab_config, tab_builder = st.tabs(["Mission & Battery Config", "Mission Builder"])

PHASE_TYPES = [
    "IDLE", "TAKEOFF", "CLIMB", "CRUISE", "HOVER",
    "DESCEND", "LAND", "PAYLOAD_OPS", "EMERGENCY",
    "VTOL_TRANSITION", "VTOL_HOVER", "FW_CRUISE", "FW_CLIMB", "FW_DESCEND",
]

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — MISSION & BATTERY CONFIG
# ══════════════════════════════════════════════════════════════════════════════
with tab_config:
    cfg = st.session_state.get("config", load_config())

    col_left, col_right = st.columns([1, 1])

    with col_left:
        st.subheader("Mission & UAV")

        mission_ids = sorted(db.missions.keys())
        if not mission_ids:
            st.error("No missions found in database.")
            mission_ids = []
            sel_mission_id = None
            mission = None
            uav = None
            sel_uav_id = None
            ambient_temp = 25
            temp_sweep = []
        else:
            default_mission = cfg.get("mission_id", mission_ids[0])
            default_mission_idx = mission_ids.index(default_mission) if default_mission in mission_ids else 0
            sel_mission_id = st.selectbox("Mission", mission_ids, index=default_mission_idx, key="cfg_mission")
            mission = db.missions[sel_mission_id]

            st.write(f"*{mission.mission_name}* — "
                     f"{mission.total_duration_s/60:.1f} min, "
                     f"{mission.total_distance_m:.0f} m, "
                     f"{len(mission.phases)} phases")

            uav_ids = sorted(db.uav_configs.keys())
            if not uav_ids:
                st.error("No UAV configurations found in database.")
                uav = None
                sel_uav_id = None
            else:
                default_uav = cfg.get("uav_id", uav_ids[0])
                default_uav_idx = uav_ids.index(default_uav) if default_uav in uav_ids else 0
                sel_uav_id = st.selectbox("UAV Configuration", uav_ids, index=default_uav_idx, key="cfg_uav")
                uav = db.uav_configs[sel_uav_id]
                st.write(f"*{uav.name}* — {uav.total_weight_g():.0f} g, {len(uav.equipment_list)} equipment items")

            default_temp = cfg.get("ambient_temp_c", 25.0)
            ambient_temp = st.slider("Ambient temperature (°C)", -20, 55, int(default_temp), key="cfg_temp")

            st.write("**Temperature sweep range**")
            sweep_existing = cfg.get("temp_sweep", list(range(-10, 46, 5)))
            col_sw1, col_sw2, col_sw3 = st.columns(3)
            with col_sw1:
                sweep_min = st.number_input("Min °C", -30, 30,
                                            int(min(sweep_existing)) if sweep_existing else -10, key="cfg_sw_min")
            with col_sw2:
                sweep_max = st.number_input("Max °C", 10, 60,
                                            int(max(sweep_existing)) if sweep_existing else 45, key="cfg_sw_max")
            with col_sw3:
                sweep_step = st.number_input("Step °C", 1, 15, 5, key="cfg_sw_step")
            temp_sweep = list(range(int(sweep_min), int(sweep_max) + 1, int(sweep_step)))
            st.caption(f"Sweep: {temp_sweep}")

    with col_right:
        st.subheader("Battery Selection")

        all_packs = db.packs
        all_pack_ids = sorted(all_packs.keys())
        default_sel = [pid for pid in cfg.get("selected_batteries", []) if pid in all_packs]

        # Initialise session state key so multiselect respects saved config on first load
        if "cfg_batteries" not in st.session_state:
            st.session_state["cfg_batteries"] = default_sel

        browser_sel = st.session_state.get("browser_selected_packs", [])
        if browser_sel:
            col_import, col_info = st.columns([1, 2])
            with col_import:
                if st.button(f"⬅ Import {len(browser_sel)} pack(s) from Battery Browser",
                             key="cfg_import_browser"):
                    current = st.session_state.get("cfg_batteries", [])
                    merged = list(dict.fromkeys(current + browser_sel))
                    st.session_state["cfg_batteries"] = merged
                    st.session_state["browser_selected_packs"] = []
                    st.success(f"Imported {len(browser_sel)} pack(s).")
                    st.rerun()
            with col_info:
                st.caption(f"Battery Browser has: {', '.join(p[:25] for p in browser_sel[:5])}"
                           + (f"… +{len(browser_sel)-5} more" if len(browser_sel) > 5 else ""))
        else:
            st.info("No packs selected in Battery Browser. Go to **Battery Browser** to filter and pick packs.",
                    icon="ℹ️")

        selected_batteries = st.multiselect(
            "Select battery packs", all_pack_ids,
            key="cfg_batteries",
            help="Select all packs you want to simulate and compare.",
        )

        with st.expander("Quick filter by chemistry / energy", expanded=False):
            filter_chem = st.multiselect(
                "Filter by chemistry",
                sorted(set(p.chemistry_id for p in all_packs.values())),
                key="cfg_filter_chem",
            )
            filter_min_wh = st.number_input("Min energy (Wh)", 0, 50000, 0, key="cfg_filter_min_wh")
            filter_max_wh = st.number_input("Max energy (Wh)", 0, 50000, 50000, key="cfg_filter_max_wh")
            filter_series = st.multiselect(
                "Filter by S count",
                sorted(set(p.cells_series for p in all_packs.values())),
                key="cfg_filter_s",
            )
            candidates = [
                pid for pid, p in all_packs.items()
                if (not filter_chem or p.chemistry_id in filter_chem)
                and filter_min_wh <= p.pack_energy_wh <= filter_max_wh
                and (not filter_series or p.cells_series in filter_series)
            ]
            st.caption(f"{len(candidates)} packs match filter")
            if st.button("Add filtered packs to selection", key="cfg_add_filter"):
                combined = list(dict.fromkeys(selected_batteries + candidates))
                st.session_state["cfg_batteries"] = combined
                st.rerun()

        if selected_batteries:
            sel_rows = []
            for pid in selected_batteries:
                p = all_packs.get(pid)
                if p:
                    sel_rows.append({
                        "ID":         p.battery_id,
                        "Chemistry":  p.chemistry_id,
                        "Energy (Wh)":p.pack_energy_wh,
                        "Weight (g)": p.pack_weight_g,
                        "Wh/kg":      p.specific_energy_wh_kg,
                        "Config":     f"{p.cells_series}S{p.cells_parallel}P",
                    })
            st.dataframe(pd.DataFrame(sel_rows), use_container_width=True, hide_index=True, height=220)
        else:
            st.info("No batteries selected yet.")

    # Mission detail + save — only if we have valid mission/uav
    if mission and uav:
        st.divider()

        with st.expander("Mission Phase Breakdown", expanded=True):
            phase_rows = []
            cumulative_wh = 0.0
            for ph in mission.phases:
                pw = ph.effective_power_w(uav)
                wh = ph.energy_wh(uav)
                cumulative_wh += wh
                phase_rows.append({
                    "#":             ph.phase_seq,
                    "Phase Name":    ph.phase_name,
                    "Type":          ph.phase_type,
                    "Duration (s)":  ph.duration_s,
                    "Power (W)":     round(pw, 0),
                    "Energy (Wh)":   round(wh, 2),
                    "Cumulative Wh": round(cumulative_wh, 2),
                    "Distance (m)":  ph.distance_m,
                    "Altitude (m)":  ph.altitude_m,
                    "Override W":    ph.power_override_w or "",
                })

            def phase_row_style(row):
                color = PHASE_COLORS.get(row["Type"], "#FFFFFF")
                return [f"background-color: {color}"] * len(row)

            st.dataframe(
                pd.DataFrame(phase_rows).style.apply(phase_row_style, axis=1).format({
                    "Power (W)": "{:.0f}", "Energy (Wh)": "{:.2f}", "Cumulative Wh": "{:.2f}"
                }),
                use_container_width=True, hide_index=True
            )

            col_total1, col_total2, col_total3 = st.columns(3)
            col_total1.metric("Total duration", f"{mission.total_duration_s/60:.1f} min")
            col_total2.metric("Total energy (UAV demand)", f"{mission.total_energy_wh(uav):.1f} Wh")
            col_total3.metric("Total distance", f"{mission.total_distance_m:.0f} m")

            fig_pw = plot_phase_power(mission, uav)
            st.pyplot(fig_pw, use_container_width=True)
            plt.close(fig_pw)

        with st.expander("UAV Equipment List", expanded=False):
            equip_rows = []
            for eq, qty in uav.equipment_list:
                equip_rows.append({
                    "ID":        eq.equip_id,
                    "Category":  eq.category,
                    "Qty":       qty,
                    "Hover W":   eq.hover_power_w * qty,
                    "Cruise W":  eq.cruise_power_w * qty,
                    "Max W":     eq.max_power_w * qty,
                    "Weight (g)":eq.weight_g * qty,
                })
            st.dataframe(pd.DataFrame(equip_rows), use_container_width=True, hide_index=True)
            st.metric("Total UAV weight (excl. battery)", f"{uav.total_weight_g():.0f} g")

        st.divider()
        st.subheader("Save Configuration")

        if not selected_batteries:
            st.warning("Select at least one battery pack before saving.")
        else:
            new_cfg = {
                "mission_id":          sel_mission_id,
                "uav_id":              sel_uav_id,
                "ambient_temp_c":      float(ambient_temp),
                "temp_sweep":          temp_sweep,
                "selected_batteries":  selected_batteries,
                "custom_equipment":    {},
                "battery_combination": None,
                "combined_pack_id":    None,
            }
            col_save, col_preview = st.columns([1, 2])
            with col_save:
                if st.button("💾 Save analysis_config.json", type="primary"):
                    save_config(new_cfg)
                    st.session_state["config"] = new_cfg
                    st.session_state["sim_results"] = {}
                    st.session_state["sweep_results"] = {}
                    st.success("Configuration saved!")
            with col_preview:
                import json
                st.code(json.dumps(new_cfg, indent=2), language="json")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — MISSION BUILDER
# ══════════════════════════════════════════════════════════════════════════════
with tab_builder:
    st.subheader("Mission Builder")
    st.caption("Create new missions or edit existing ones. Add, modify, and reorder flight phases.")

    mode = st.radio("Mode", ["Create new mission", "Edit existing mission"],
                    horizontal=True, key="mb_mode")

    all_mission_ids = sorted(db.missions.keys())

    # ── Working context ───────────────────────────────────────────────────────
    if mode == "Edit existing mission":
        if not all_mission_ids:
            st.warning("No missions in database yet. Switch to 'Create new mission'.")
            working_mission_id   = ""
            working_mission_name = ""
            working_uav_id       = ""
            edit_mission         = None
        else:
            edit_mid = st.selectbox("Select mission to edit", all_mission_ids, key="mb_edit_sel")
            edit_mission = db.missions[edit_mid]
            working_mission_id   = edit_mid
            working_mission_name = edit_mission.mission_name

            st.write(f"**{edit_mission.mission_name}** — {len(edit_mission.phases)} phases, "
                     f"{edit_mission.total_duration_s/60:.1f} min")

            # Current phases table
            exist_rows = []
            for ph in sorted(edit_mission.phases, key=lambda p: p.phase_seq):
                est_w = ph.power_override_w or "auto"
                exist_rows.append({
                    "Seq":          ph.phase_seq,
                    "Phase Name":   ph.phase_name,
                    "Type":         ph.phase_type,
                    "Duration (s)": ph.duration_s,
                    "Distance (m)": ph.distance_m,
                    "Altitude (m)": ph.altitude_m,
                    "Airspeed m/s": ph.airspeed_ms,
                    "Override W":   est_w,
                    "Notes":        ph.notes,
                })
            st.dataframe(pd.DataFrame(exist_rows), use_container_width=True, hide_index=True)

            uav_id_options = sorted(db.uav_configs.keys())
            working_uav_id = st.selectbox(
                "UAV config (for power preview)",
                uav_id_options if uav_id_options else ["—"],
                key="mb_edit_uav",
            )
            next_seq = max((ph.phase_seq for ph in edit_mission.phases), default=0) + 1

    else:  # Create new mission
        edit_mission = None
        col_m1, col_m2, col_m3 = st.columns(3)
        with col_m1:
            working_mission_id = st.text_input("Mission ID*", key="mb_new_id",
                                               help="Unique ID e.g. SURVEY_ROUTE_A_V1")
        with col_m2:
            working_mission_name = st.text_input("Mission Name*", key="mb_new_name",
                                                  help="e.g. Survey Route Alpha")
        with col_m3:
            uav_id_options = sorted(db.uav_configs.keys())
            working_uav_id = st.selectbox(
                "Default UAV config",
                uav_id_options if uav_id_options else ["—"],
                key="mb_new_uav",
            )

        next_seq = 1
        staged_key = f"mb_staged_{working_mission_id}" if working_mission_id else "mb_staged_"
        staged = st.session_state.get(staged_key, [])

        if working_mission_id in db.missions:
            st.warning(f"Mission ID `{working_mission_id}` already exists — switch to 'Edit existing'.")
            working_mission_id = ""

        if working_mission_id and staged:
            st.write(f"**Staged phases ({len(staged)}):**")
            st.dataframe(pd.DataFrame(staged), use_container_width=True, hide_index=True)
            next_seq = max(r["Seq"] for r in staged) + 1
            if st.button("🗑️ Clear staged phases", key="mb_clear_staged"):
                st.session_state[staged_key] = []
                st.rerun()

    st.divider()

    # ── Add Phase Form ────────────────────────────────────────────────────────
    can_add = bool(working_mission_id)
    if not can_add:
        st.info("Enter a Mission ID above to start adding phases.")
    else:
        st.subheader("Add Phase")

        col_p1, col_p2, col_p3 = st.columns(3)
        with col_p1:
            ph_seq      = st.number_input("Sequence #", 1, 999, next_seq, key="mb_ph_seq")
            ph_name     = st.text_input("Phase Name*", key="mb_ph_name",
                                        help="e.g. Initial Climb, Waypoint Cruise")
            ph_type     = st.selectbox("Phase Type*", PHASE_TYPES, key="mb_ph_type")
        with col_p2:
            ph_duration = st.number_input("Duration (s)*", 0.0, 86400.0, 60.0, 1.0, key="mb_ph_dur")
            ph_distance = st.number_input("Distance (m)", 0.0, 500000.0, 0.0, 10.0, key="mb_ph_dist")
            ph_altitude = st.number_input("Altitude (m AGL)", 0.0, 5000.0, 0.0, 1.0, key="mb_ph_alt")
        with col_p3:
            ph_airspeed = st.number_input("Airspeed (m/s)", 0.0, 200.0, 0.0, 0.5, key="mb_ph_spd")
            ph_override = st.number_input("Power override (W, 0 = auto)", 0.0, 100000.0, 0.0, 10.0,
                                          key="mb_ph_ovr")
            ph_notes    = st.text_input("Notes", key="mb_ph_notes")

        # Live power/energy estimate
        if working_uav_id in db.uav_configs:
            est_w  = ph_override if ph_override > 0 else db.uav_configs[working_uav_id].phase_power_w(ph_type)
            est_wh = est_w * ph_duration / 3600
            st.caption(f"Estimated: **{est_w:.0f} W** | **{est_wh:.2f} Wh** for this phase")

        col_add, col_save_new = st.columns([1, 2])

        with col_add:
            if st.button("➕ Add Phase", type="primary", key="mb_add_phase"):
                if not ph_name:
                    st.error("Phase Name is required.")
                elif ph_duration <= 0:
                    st.error("Duration must be > 0.")
                else:
                    phase_row = {
                        "Seq": ph_seq, "Phase Name": ph_name, "Type": ph_type,
                        "Duration (s)": ph_duration, "Distance (m)": ph_distance,
                        "Altitude (m)": ph_altitude, "Airspeed m/s": ph_airspeed,
                        "Override W": ph_override if ph_override > 0 else None,
                        "Notes": ph_notes,
                    }

                    if mode == "Create new mission":
                        staged_key = f"mb_staged_{working_mission_id}"
                        if staged_key not in st.session_state:
                            st.session_state[staged_key] = []
                        st.session_state[staged_key].append(phase_row)
                        st.success(f"Phase '{ph_name}' staged (seq {ph_seq}).")
                        st.rerun()
                    else:
                        # Write directly to sheet
                        try:
                            from openpyxl import load_workbook
                            wb = load_workbook(DB_PATH)
                            ws = wb["Mission_Profiles"]
                            ws.append([
                                working_mission_id, working_mission_name,
                                working_uav_id, ph_seq, ph_name, ph_type,
                                ph_duration, ph_distance, ph_altitude,
                                ph_airspeed,
                                ph_override if ph_override > 0 else None,
                                ph_notes,
                            ])
                            wb.save(DB_PATH)
                            reload_db()
                            st.success(f"Phase '{ph_name}' added to `{working_mission_id}`.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Failed to save phase: {e}")

        with col_save_new:
            if mode == "Create new mission":
                staged_key = f"mb_staged_{working_mission_id}"
                staged = st.session_state.get(staged_key, [])
                if staged:
                    if st.button(f"💾 Save mission '{working_mission_id}' ({len(staged)} phases)",
                                 type="primary", key="mb_save_mission"):
                        if not working_mission_name:
                            st.error("Mission Name is required before saving.")
                        else:
                            try:
                                from openpyxl import load_workbook
                                wb = load_workbook(DB_PATH)
                                ws = wb["Mission_Profiles"]
                                for row in staged:
                                    ws.append([
                                        working_mission_id, working_mission_name,
                                        working_uav_id,
                                        row["Seq"], row["Phase Name"], row["Type"],
                                        row["Duration (s)"], row["Distance (m)"],
                                        row["Altitude (m)"], row["Airspeed m/s"],
                                        row["Override W"], row["Notes"],
                                    ])
                                wb.save(DB_PATH)
                                reload_db()
                                st.session_state[staged_key] = []
                                st.success(f"Mission `{working_mission_id}` saved with {len(staged)} phases!")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Failed to save mission: {e}")

    # ── Edit existing phase ───────────────────────────────────────────────────
    if mode == "Edit existing mission" and edit_mission and edit_mission.phases:
        st.divider()
        with st.expander("✏️ Edit an Existing Phase", expanded=False):
            phase_options = {
                f"#{ph.phase_seq} — {ph.phase_name} ({ph.phase_type})": ph
                for ph in sorted(edit_mission.phases, key=lambda p: p.phase_seq)
            }
            sel_phase_label = st.selectbox("Select phase to edit", list(phase_options.keys()),
                                           key="mb_edit_phase_sel")
            sel_ph = phase_options[sel_phase_label]

            col_e1, col_e2, col_e3 = st.columns(3)
            with col_e1:
                e_name = st.text_input("Phase Name", value=sel_ph.phase_name, key="mb_e_name")
                e_type = st.selectbox("Phase Type", PHASE_TYPES,
                                      index=PHASE_TYPES.index(sel_ph.phase_type)
                                            if sel_ph.phase_type in PHASE_TYPES else 0,
                                      key="mb_e_type")
            with col_e2:
                e_dur  = st.number_input("Duration (s)",  0.0, 86400.0,  float(sel_ph.duration_s),  1.0, key="mb_e_dur")
                e_dist = st.number_input("Distance (m)",  0.0, 500000.0, float(sel_ph.distance_m),  10.0, key="mb_e_dist")
                e_alt  = st.number_input("Altitude (m)",  0.0, 5000.0,   float(sel_ph.altitude_m),  1.0, key="mb_e_alt")
            with col_e3:
                e_spd   = st.number_input("Airspeed (m/s)", 0.0, 200.0, float(sel_ph.airspeed_ms), 0.5, key="mb_e_spd")
                e_ovr   = st.number_input("Power override (W, 0 = auto)", 0.0, 100000.0,
                                          float(sel_ph.power_override_w or 0), 10.0, key="mb_e_ovr")
                e_notes = st.text_input("Notes", value=sel_ph.notes, key="mb_e_notes")

            if st.button("💾 Update Phase", type="primary", key="mb_update_phase"):
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(DB_PATH)
                    ws = wb["Mission_Profiles"]
                    updated = False
                    for row in ws.iter_rows(min_row=4):
                        if (row[0].value == working_mission_id and
                                row[3].value == sel_ph.phase_seq):
                            row[4].value  = e_name
                            row[5].value  = e_type
                            row[6].value  = e_dur
                            row[7].value  = e_dist
                            row[8].value  = e_alt
                            row[9].value  = e_spd
                            row[10].value = e_ovr if e_ovr > 0 else None
                            row[11].value = e_notes
                            updated = True
                            break
                    if updated:
                        wb.save(DB_PATH)
                        reload_db()
                        st.success(f"Phase #{sel_ph.phase_seq} '{e_name}' updated.")
                        st.rerun()
                    else:
                        st.error("Phase row not found in workbook.")
                except Exception as e:
                    st.error(f"Failed to update: {e}")

    # ── Delete an existing phase ──────────────────────────────────────────────
    if mode == "Edit existing mission" and edit_mission and edit_mission.phases:
        with st.expander("🗑️ Delete a Phase", expanded=False):
            st.warning("This permanently removes the phase row from the database.")
            del_phase_options = {
                f"#{ph.phase_seq} — {ph.phase_name} ({ph.phase_type})": ph
                for ph in sorted(edit_mission.phases, key=lambda p: p.phase_seq)
            }
            del_phase_label = st.selectbox("Select phase to delete", list(del_phase_options.keys()),
                                           key="mb_del_phase_sel")
            del_ph = del_phase_options[del_phase_label]
            st.write(f"Will delete: **#{del_ph.phase_seq} — {del_ph.phase_name}** "
                     f"({del_ph.phase_type}, {del_ph.duration_s} s)")
            if st.button("🗑️ Confirm Delete Phase", type="primary", key="mb_del_phase_confirm"):
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(DB_PATH)
                    ws = wb["Mission_Profiles"]
                    row_to_delete = None
                    for row in ws.iter_rows(min_row=4):
                        if (row[0].value == working_mission_id and
                                row[3].value == del_ph.phase_seq):
                            row_to_delete = row[0].row
                            break
                    if row_to_delete:
                        ws.delete_rows(row_to_delete)
                        wb.save(DB_PATH)
                        reload_db()
                        st.success(f"Phase #{del_ph.phase_seq} '{del_ph.phase_name}' deleted.")
                        st.rerun()
                    else:
                        st.error("Phase row not found in workbook.")
                except Exception as e:
                    st.error(f"Failed to delete phase: {e}")

    # ── All missions summary ──────────────────────────────────────────────────
    st.divider()
    st.subheader("All Missions")
    if all_mission_ids:
        sum_rows = []
        for mid, m in db.missions.items():
            sum_rows.append({
                "Mission ID":    mid,
                "Name":          m.mission_name,
                "Phases":        len(m.phases),
                "Duration (min)":round(m.total_duration_s / 60, 1),
                "Distance (m)":  round(m.total_distance_m, 0),
            })
        st.dataframe(pd.DataFrame(sum_rows).sort_values("Mission ID"),
                     use_container_width=True, hide_index=True)
    else:
        st.info("No missions in database yet.")
