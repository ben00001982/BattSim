"""
ui/pages/4_Equipment_Builder.py
Add and edit equipment items, configure UAV equipment profiles, and manage UAV configurations.
"""
import sys
from pathlib import Path
_PROJECT_ROOT = Path(__file__).parent.parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

import streamlit as st
import pandas as pd
from openpyxl import load_workbook

st.set_page_config(page_title="Equipment Builder", page_icon="🛠️", layout="wide")

from ui.components.db_helpers import load_db, reload_db
from ui.config import DB_PATH

st.title("🛠️ Equipment & Power Profile Builder")
st.caption("Add and edit equipment items, configure UAV power profiles, and manage UAV configurations.")

db = load_db()

tab_equip, tab_uav = st.tabs(["Equipment Database", "UAV Configurations"])

# ── Tab 1: Equipment Database ─────────────────────────────────────────────────
with tab_equip:
    st.subheader("Equipment Database")

    # Display existing equipment
    if db.equipment:
        equip_rows = []
        for eq in db.equipment.values():
            equip_rows.append({
                "ID":           eq.equip_id,
                "Category":     eq.category,
                "Manufacturer": eq.manufacturer,
                "Model":        eq.model,
                "Idle W":       eq.idle_power_w,
                "Hover W":      eq.hover_power_w,
                "Climb W":      eq.climb_power_w,
                "Cruise W":     eq.cruise_power_w,
                "Max W":        eq.max_power_w,
                "Weight (g)":   eq.weight_g,
                "Eff %":        eq.efficiency_pct,
                "Active":       "Yes" if eq.active else "No",
                "Notes":        eq.notes,
            })
        equip_df = pd.DataFrame(equip_rows)
        st.dataframe(equip_df, use_container_width=True, hide_index=True, height=280)
    else:
        st.info("No equipment found in database.")

    st.divider()

    with st.expander("➕ Add New Equipment Item", expanded=False):
        st.caption("Fill in all fields and click Save to add the item to Equipment_DB sheet.")

        col_a, col_b, col_c = st.columns(3)
        with col_a:
            new_equip_id   = st.text_input("Equipment ID*", key="eq_id",
                                            help="Unique ID, e.g. MOTORS_DJI_E5000")
            new_category   = st.selectbox("Category*", [
                "MOTORS_PROPS", "FLIGHT_CONTROLLER", "ESC", "GIMBAL", "PAYLOAD",
                "TELEMETRY", "GPS", "LIGHTS", "CAMERA", "SENSORS", "OTHER"
            ], key="eq_cat")
            new_mfr        = st.text_input("Manufacturer", key="eq_mfr")
            new_model      = st.text_input("Model", key="eq_model")
        with col_b:
            new_nom_v      = st.number_input("Nom voltage (V)", 0.0, 100.0, 0.0, 0.1, key="eq_nom_v")
            new_nom_a      = st.number_input("Nom current (A)", 0.0, 1000.0, 0.0, 0.5, key="eq_nom_a")
            new_idle_w     = st.number_input("Idle power (W)", 0.0, 5000.0, 0.0, 1.0, key="eq_idle_w")
            new_hover_w    = st.number_input("Hover power (W)*", 0.0, 50000.0, 0.0, 10.0, key="eq_hover_w")
            new_climb_w    = st.number_input("Climb power (W)", 0.0, 50000.0, 0.0, 10.0, key="eq_climb_w")
        with col_c:
            new_cruise_w   = st.number_input("Cruise power (W)*", 0.0, 50000.0, 0.0, 10.0, key="eq_cruise_w")
            new_max_w      = st.number_input("Max power (W)", 0.0, 100000.0, 0.0, 10.0, key="eq_max_w")
            new_weight_g   = st.number_input("Weight (g)", 0.0, 50000.0, 0.0, 1.0, key="eq_weight_g")
            new_eff_pct    = st.number_input("Efficiency (%)", 0.0, 100.0, 95.0, 0.5, key="eq_eff")
            new_duty_pct   = st.number_input("Duty cycle (%)", 0.0, 100.0, 100.0, 1.0, key="eq_duty")

        new_notes  = st.text_input("Notes", key="eq_notes")
        new_active = st.checkbox("Active", value=True, key="eq_active")

        if st.button("💾 Save Equipment to Database", type="primary", key="eq_save"):
            if not new_equip_id:
                st.error("Equipment ID is required.")
            elif new_equip_id in db.equipment:
                st.error(f"Equipment ID `{new_equip_id}` already exists.")
            else:
                try:
                    wb = load_workbook(DB_PATH)
                    ws = wb["Equipment_DB"]
                    ws.append([
                        new_equip_id, new_category, new_mfr, new_model,
                        new_nom_v, new_nom_a, new_idle_w, new_hover_w,
                        new_climb_w, new_cruise_w, new_max_w, new_weight_g,
                        new_eff_pct, new_duty_pct, new_notes,
                        "YES" if new_active else "NO",
                    ])
                    wb.save(DB_PATH)
                    reload_db()
                    st.success(f"Equipment `{new_equip_id}` saved to database!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Failed to save: {e}")

    # Edit existing equipment power values
    with st.expander("✏️ Edit Equipment Power Values", expanded=False):
        st.caption("Update hover/cruise/max power for an existing equipment item.")
        all_eq_ids = sorted(db.equipment.keys())
        if not all_eq_ids:
            st.info("No equipment to edit.")
        else:
            edit_eq_id = st.selectbox("Select equipment to edit", all_eq_ids, key="eq_edit_sel")
            eq_obj = db.equipment[edit_eq_id]

            col_e1, col_e2, col_e3 = st.columns(3)
            with col_e1:
                edit_idle_w  = st.number_input("Idle W",  0.0, 50000.0, float(eq_obj.idle_power_w),  1.0, key="eq_edit_idle")
                edit_hover_w = st.number_input("Hover W", 0.0, 50000.0, float(eq_obj.hover_power_w), 1.0, key="eq_edit_hover")
            with col_e2:
                edit_climb_w  = st.number_input("Climb W",  0.0, 50000.0, float(eq_obj.climb_power_w),  1.0, key="eq_edit_climb")
                edit_cruise_w = st.number_input("Cruise W", 0.0, 50000.0, float(eq_obj.cruise_power_w), 1.0, key="eq_edit_cruise")
            with col_e3:
                edit_max_w    = st.number_input("Max W",    0.0, 100000.0, float(eq_obj.max_power_w),   1.0, key="eq_edit_max")
                edit_weight   = st.number_input("Weight (g)", 0.0, 50000.0, float(eq_obj.weight_g),    1.0, key="eq_edit_weight")

            if st.button("💾 Update Equipment", key="eq_edit_save"):
                try:
                    wb = load_workbook(DB_PATH)
                    ws = wb["Equipment_DB"]
                    # Find and update row
                    for row in ws.iter_rows(min_row=3):
                        if row[0].value == edit_eq_id:
                            row[6].value  = edit_idle_w    # col G
                            row[7].value  = edit_hover_w   # col H
                            row[8].value  = edit_climb_w   # col I
                            row[9].value  = edit_cruise_w  # col J
                            row[10].value = edit_max_w     # col K
                            row[11].value = edit_weight    # col L
                            break
                    wb.save(DB_PATH)
                    reload_db()
                    st.success(f"Equipment `{edit_eq_id}` updated!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Failed to update: {e}")

# ── Tab 2: UAV Configurations ─────────────────────────────────────────────────
with tab_uav:
    st.subheader("UAV Configurations")

    # Show existing UAV configs
    if db.uav_configs:
        uav_summary = []
        for uid, uav in db.uav_configs.items():
            uav_summary.append({
                "UAV ID":       uid,
                "Name":         uav.name,
                "Equipment":    len(uav.equipment_list),
                "Total weight (g)": round(uav.total_weight_g(), 0),
                "Hover W (est.)": round(uav.phase_power_w("HOVER"), 0),
                "Cruise W (est.)": round(uav.phase_power_w("CRUISE"), 0),
            })
        st.dataframe(pd.DataFrame(uav_summary), use_container_width=True, hide_index=True)
    else:
        st.info("No UAV configurations found.")

    st.divider()

    # View details of a selected UAV
    with st.expander("🔍 View UAV Equipment List", expanded=False):
        uav_ids = sorted(db.uav_configs.keys())
        if uav_ids:
            view_uav_id = st.selectbox("Select UAV", uav_ids, key="view_uav_sel")
            view_uav = db.uav_configs[view_uav_id]
            view_rows = []
            for eq, qty in view_uav.equipment_list:
                view_rows.append({
                    "Equipment ID": eq.equip_id,
                    "Category":     eq.category,
                    "Manufacturer": eq.manufacturer,
                    "Model":        eq.model,
                    "Qty":          qty,
                    "Hover W (total)":  round(eq.hover_power_w * qty, 1),
                    "Cruise W (total)": round(eq.cruise_power_w * qty, 1),
                    "Max W (total)":    round(eq.max_power_w * qty, 1),
                    "Weight g (total)": round(eq.weight_g * qty, 1),
                })
            st.dataframe(pd.DataFrame(view_rows), use_container_width=True, hide_index=True)
            c1, c2, c3 = st.columns(3)
            c1.metric("Total weight", f"{view_uav.total_weight_g():.0f} g")
            c2.metric("Hover power", f"{view_uav.phase_power_w('HOVER'):.0f} W")
            c3.metric("Cruise power", f"{view_uav.phase_power_w('CRUISE'):.0f} W")

    # Create new UAV configuration
    with st.expander("➕ Create New UAV Configuration", expanded=False):
        st.caption("Define a new UAV by adding equipment items. Saves to UAV_Configurations sheet.")

        new_uav_id   = st.text_input("UAV Config ID*", key="new_uav_id",
                                      help="Unique ID e.g. M350_SURVEY_V1")
        new_uav_name = st.text_input("UAV Name", key="new_uav_name",
                                     help="Descriptive name e.g. DJI M350 Survey Config")

        st.write("**Add equipment items:**")
        all_eq_ids = sorted(db.equipment.keys())
        if not all_eq_ids:
            st.warning("No equipment in database. Add equipment first.")
        else:
            # Use session state to manage equipment list being built
            if "new_uav_items" not in st.session_state:
                st.session_state["new_uav_items"] = []

            col_eq1, col_eq2, col_eq3 = st.columns([3, 1, 1])
            with col_eq1:
                add_eq_id  = st.selectbox("Equipment", all_eq_ids, key="new_uav_eq_sel")
            with col_eq2:
                add_eq_qty = st.number_input("Qty", 1, 50, 1, key="new_uav_eq_qty")
            with col_eq3:
                st.write("")
                st.write("")
                if st.button("➕ Add", key="new_uav_add_eq"):
                    existing_ids = [item[0] for item in st.session_state["new_uav_items"]]
                    if add_eq_id in existing_ids:
                        idx = existing_ids.index(add_eq_id)
                        st.session_state["new_uav_items"][idx] = (add_eq_id, add_eq_qty)
                    else:
                        st.session_state["new_uav_items"].append((add_eq_id, add_eq_qty))
                    st.rerun()

            # Show current items
            if st.session_state["new_uav_items"]:
                preview_rows = []
                total_w = 0.0
                total_hover_w = 0.0
                total_cruise_w = 0.0
                for eid, qty in st.session_state["new_uav_items"]:
                    eq = db.equipment.get(eid)
                    if eq:
                        preview_rows.append({
                            "Equipment ID": eid,
                            "Qty": qty,
                            "Hover W": eq.hover_power_w * qty,
                            "Cruise W": eq.cruise_power_w * qty,
                            "Weight g": eq.weight_g * qty,
                        })
                        total_w += eq.weight_g * qty
                        total_hover_w += eq.hover_power_w * qty
                        total_cruise_w += eq.cruise_power_w * qty
                st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True)
                c1, c2, c3 = st.columns(3)
                c1.metric("Total weight", f"{total_w:.0f} g")
                c2.metric("Total hover W", f"{total_hover_w:.0f} W")
                c3.metric("Total cruise W", f"{total_cruise_w:.0f} W")

                col_clr, col_sav = st.columns([1, 2])
                with col_clr:
                    if st.button("🗑️ Clear list", key="new_uav_clear"):
                        st.session_state["new_uav_items"] = []
                        st.rerun()
                with col_sav:
                    if st.button("💾 Save UAV Configuration", type="primary", key="new_uav_save"):
                        if not new_uav_id:
                            st.error("UAV Config ID is required.")
                        elif new_uav_id in db.uav_configs:
                            st.error(f"UAV ID `{new_uav_id}` already exists.")
                        else:
                            try:
                                wb  = load_workbook(DB_PATH)
                                ws  = wb["UAV_Configurations"]
                                # Write each equipment row
                                for eid, qty in st.session_state["new_uav_items"]:
                                    ws.append([new_uav_id, new_uav_name or new_uav_id, eid, qty])
                                wb.save(DB_PATH)
                                reload_db()
                                st.session_state["new_uav_items"] = []
                                st.success(f"UAV configuration `{new_uav_id}` saved!")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Failed to save: {e}")

    # Edit existing UAV - add equipment
    with st.expander("✏️ Edit Existing UAV — Add Equipment Item", expanded=False):
        st.caption("Append a new equipment item to an existing UAV configuration.")
        uav_ids = sorted(db.uav_configs.keys())
        if not uav_ids:
            st.info("No UAV configurations to edit.")
        else:
            edit_uav_id = st.selectbox("Select UAV to edit", uav_ids, key="edit_uav_sel")
            all_eq_ids  = sorted(db.equipment.keys())

            col_ea, col_eb, col_ec = st.columns([3, 1, 1])
            with col_ea:
                edit_add_eq = st.selectbox("Equipment to add", all_eq_ids, key="edit_uav_eq")
            with col_eb:
                edit_add_qty = st.number_input("Qty", 1, 50, 1, key="edit_uav_qty")
            with col_ec:
                st.write("")
                st.write("")
                if st.button("➕ Append to UAV", key="edit_uav_append"):
                    try:
                        wb  = load_workbook(DB_PATH)
                        ws  = wb["UAV_Configurations"]
                        uav_obj = db.uav_configs[edit_uav_id]
                        ws.append([edit_uav_id, uav_obj.name, edit_add_eq, edit_add_qty])
                        wb.save(DB_PATH)
                        reload_db()
                        st.success(f"Added `{edit_add_eq}` ×{edit_add_qty} to `{edit_uav_id}`.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Failed to update: {e}")

    # Remove equipment from UAV
    with st.expander("🗑️ Remove Equipment from UAV", expanded=False):
        st.caption("Permanently removes one equipment row from a UAV configuration.")
        uav_ids_rm = sorted(db.uav_configs.keys())
        if not uav_ids_rm:
            st.info("No UAV configurations to edit.")
        else:
            rm_uav_id = st.selectbox("Select UAV", uav_ids_rm, key="rm_uav_sel")
            rm_uav    = db.uav_configs[rm_uav_id]
            if not rm_uav.equipment_list:
                st.info("This UAV has no equipment items.")
            else:
                rm_eq_options = {
                    f"{eq.equip_id} ×{qty} ({eq.category})": eq.equip_id
                    for eq, qty in rm_uav.equipment_list
                }
                rm_label = st.selectbox("Equipment to remove", list(rm_eq_options.keys()),
                                        key="rm_uav_eq_sel")
                rm_eq_id = rm_eq_options[rm_label]
                st.warning(f"Will remove `{rm_eq_id}` from `{rm_uav_id}`.")
                if st.button("🗑️ Confirm Remove", type="primary", key="rm_uav_confirm"):
                    try:
                        wb = load_workbook(DB_PATH)
                        ws = wb["UAV_Configurations"]
                        row_to_delete = None
                        for row in ws.iter_rows(min_row=3):
                            if row[0].value == rm_uav_id and row[2].value == rm_eq_id:
                                row_to_delete = row[0].row
                                break
                        if row_to_delete:
                            ws.delete_rows(row_to_delete)
                            wb.save(DB_PATH)
                            reload_db()
                            st.success(f"Removed `{rm_eq_id}` from `{rm_uav_id}`.")
                            st.rerun()
                        else:
                            st.error("Row not found in workbook.")
                    except Exception as e:
                        st.error(f"Failed to remove: {e}")
