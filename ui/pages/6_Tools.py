"""
ui/pages/6_Tools.py
Tools: Pack Builder, Model Validation, and Bulk Data Upload.
"""
import sys
from pathlib import Path
_PROJECT_ROOT = Path(__file__).parent.parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import io

st.set_page_config(page_title="Tools", page_icon="🔧", layout="wide")

from ui.components.db_helpers import load_db, cells_to_df, packs_to_df, reload_db, load_config
from ui.config import ACCENT

st.title("🔧 Tools")
st.caption("Pack Builder, Model Validation, and Bulk Data Upload.")

db  = load_db()
cfg = load_config()

tab_builder, tab_validate, tab_bulk = st.tabs(
    ["Pack Builder", "Model Validation", "Bulk Data Upload"]
)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — PACK BUILDER
# ══════════════════════════════════════════════════════════════════════════════
with tab_builder:
    st.subheader("🔧 Pack Builder")
    st.caption("Assemble custom battery packs from cells or combine existing packs.")

    from batteries.builder import build_pack, combine_packs

    sub_build, sub_combine = st.tabs(["Build from Cell", "Combine Packs"])

    with sub_build:
        col_left, col_right = st.columns([1, 1])

        with col_left:
            cell_ids = sorted(db.cells.keys())
            if not cell_ids:
                st.warning("No cells found in database.")
            else:
                sel_cell_id = st.selectbox("Select cell", cell_ids, key="builder_cell")
                cell = db.cells[sel_cell_id]

                st.markdown(f"**{cell.manufacturer} {cell.model}**")
                st.write(f"Chemistry: `{cell.chemistry_id}` | Format: `{cell.cell_format}`")

                cell_info = {
                    "Nominal V": f"{cell.voltage_nominal:.2f} V",
                    "Max V":     f"{cell.voltage_max:.2f} V",
                    "Cutoff V":  f"{cell.voltage_cutoff:.2f} V",
                    "Capacity":  f"{cell.capacity_ah:.3f} Ah",
                    "Energy":    f"{cell.energy_wh:.2f} Wh",
                    "Weight":    f"{cell.weight_g:.1f} g",
                    "Sp. Energy":f"{cell.specific_energy_wh_kg:.0f} Wh/kg",
                    "Max I":     f"{cell.max_cont_discharge_a:.1f} A",
                    "IR":        f"{cell.internal_resistance_mohm:.1f} mΩ",
                    "Cycles":    str(cell.cycle_life),
                }
                st.table(pd.DataFrame({"Spec": cell_info.keys(), "Value": cell_info.values()}))

        with col_right:
            if cell_ids:
                series          = st.number_input("Cells in series (S)", 1, 28, 6, key="builder_s")
                parallel        = st.number_input("Cells in parallel (P)", 1, 50, 1, key="builder_p")
                overhead_pct    = st.slider("Weight overhead % (BMS + wiring)", 5, 30, 12, key="builder_oh")
                pack_id_input   = st.text_input("Pack ID (leave blank to auto-generate)", key="builder_id")
                pack_name_input = st.text_input("Pack name (leave blank to auto-generate)", key="builder_name")
                uav_class_input = st.text_input("UAV class", key="builder_uav")
                notes_input     = st.text_area("Notes", key="builder_notes", height=60)

        if cell_ids:
            try:
                preview = build_pack(
                    cell=cell, series=series, parallel=parallel,
                    pack_id=pack_id_input or "", name=pack_name_input or "",
                    overhead_pct=overhead_pct, uav_class=uav_class_input,
                    notes=notes_input,
                )
                st.divider()
                st.subheader("Computed Pack Specifications")
                cols = st.columns(5)
                cols[0].metric("Voltage (nom)", f"{preview.pack_voltage_nom:.2f} V")
                cols[1].metric("Capacity",      f"{preview.pack_capacity_ah:.2f} Ah")
                cols[2].metric("Energy",        f"{preview.pack_energy_wh:.1f} Wh")
                cols[3].metric("Weight",        f"{preview.pack_weight_g:.0f} g")
                cols[4].metric("Sp. Energy",    f"{preview.specific_energy_wh_kg:.0f} Wh/kg")

                cols2 = st.columns(4)
                cols2[0].metric("Config",       f"{preview.cells_series}S{preview.cells_parallel}P")
                cols2[1].metric("Max I (cont)", f"{preview.max_cont_discharge_a:.1f} A")
                cols2[2].metric("Max W (cont)", f"{preview.max_cont_discharge_w:.0f} W")
                cols2[3].metric("IR (mΩ)",      f"{preview.internal_resistance_mohm:.2f} mΩ")

                st.divider()
                already_exists = preview.battery_id in db.packs
                if already_exists:
                    st.warning(f"Pack ID `{preview.battery_id}` already exists in the database.")

                col_save, col_info = st.columns([1, 3])
                with col_save:
                    if st.button("💾 Save to Database", disabled=already_exists, key="builder_save"):
                        try:
                            db.append_custom_pack(preview)
                            reload_db()
                            st.success(f"Pack `{preview.battery_id}` saved!")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Failed to save: {e}")
                with col_info:
                    if already_exists:
                        st.info("Change the Pack ID above to save as a new entry.")
                    else:
                        st.info("Saving writes directly to battery_db.xlsx (Battery_Catalog sheet).")
            except ValueError as e:
                st.error(str(e))

    with sub_combine:
        st.subheader("Combine Existing Packs")
        st.caption("Connect packs in series (higher voltage) or parallel (higher capacity). "
                   "Add multiple instances of the same pack using the quantity field.")

        all_pack_ids = sorted(db.packs.keys())
        topology      = st.radio("Connection topology", ["series", "parallel"], horizontal=True, key="combine_topo")
        combined_id   = st.text_input("Combined Pack ID (leave blank to auto)", key="combine_id")
        combined_name = st.text_input("Combined Pack Name (leave blank to auto)", key="combine_name")

        if "combine_items" not in st.session_state:
            st.session_state["combine_items"] = []  # list of (pack_id, qty)

        col_ca, col_cb, col_cc = st.columns([3, 1, 1])
        with col_ca:
            add_comb_pid = st.selectbox("Pack to add", all_pack_ids, key="combine_add_sel")
        with col_cb:
            add_comb_qty = st.number_input("Qty", 1, 20, 1, key="combine_add_qty")
        with col_cc:
            st.write("")
            st.write("")
            if st.button("➕ Add", key="combine_add_btn"):
                st.session_state["combine_items"].append((add_comb_pid, int(add_comb_qty)))
                st.rerun()

        if st.session_state["combine_items"]:
            preview_rows = []
            for pid, qty in st.session_state["combine_items"]:
                p = db.packs.get(pid)
                if p:
                    preview_rows.append({
                        "Pack ID": pid, "Qty": qty,
                        "Chemistry": p.chemistry_id,
                        "Config": f"{p.cells_series}S{p.cells_parallel}P",
                        "Energy (Wh)": round(p.pack_energy_wh * qty, 1),
                        "Weight (g)": round(p.pack_weight_g * qty, 0),
                    })
            st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True)

            col_clr2, _ = st.columns([1, 3])
            with col_clr2:
                if st.button("🗑️ Clear list", key="combine_clear"):
                    st.session_state["combine_items"] = []
                    st.rerun()

            # Expand items by qty into flat list for combine_packs
            expanded_ids = [pid for pid, qty in st.session_state["combine_items"] for _ in range(qty)]
            if len(expanded_ids) < 2:
                st.info("Add at least 2 pack instances total to combine.")
            else:
                packs_to_comb = [db.packs[pid] for pid in expanded_ids if pid in db.packs]
                chems = set(p.chemistry_id for p in packs_to_comb)
                if len(chems) > 1:
                    st.warning(f"Mixed chemistries: {', '.join(chems)}. May produce unreliable results.")

                try:
                    combined = combine_packs(packs_to_comb, topology, combined_id, combined_name)
                except Exception as e:
                    st.error(str(e))
                    combined = None

                if combined is not None:
                    st.subheader("Combined Pack Preview")
                    c1, c2, c3, c4 = st.columns(4)
                    c1.metric("Voltage (nom)", f"{combined.pack_voltage_nom:.2f} V")
                    c2.metric("Capacity",      f"{combined.pack_capacity_ah:.2f} Ah")
                    c3.metric("Energy",        f"{combined.pack_energy_wh:.1f} Wh")
                    c4.metric("Weight",        f"{combined.pack_weight_g:.0f} g")

                    c5, c6, c7 = st.columns(3)
                    c5.metric("Sp. Energy",    f"{combined.specific_energy_wh_kg:.0f} Wh/kg")
                    c6.metric("Max I",         f"{combined.max_cont_discharge_a:.1f} A")
                    c7.metric("IR (mΩ)",       f"{combined.internal_resistance_mohm:.2f}")

                    already_c = combined.battery_id in db.packs
                    if already_c:
                        st.warning(f"Pack ID `{combined.battery_id}` already exists.")

                    if st.button("💾 Save Combined Pack", disabled=already_c, key="combine_save"):
                        try:
                            db.append_custom_pack(combined)
                            reload_db()
                            st.success(f"Combined pack `{combined.battery_id}` saved!")
                            st.session_state["combine_items"] = []
                            st.rerun()
                        except Exception as e:
                            st.error(f"Failed to save: {e}")
        else:
            st.info("Add packs above to build your combination.")

    st.divider()
    with st.expander("View Cell Catalog", expanded=False):
        cells_df = cells_to_df(db.cells)
        st.dataframe(cells_df, use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — MODEL VALIDATION
# ══════════════════════════════════════════════════════════════════════════════
with tab_validate:
    st.subheader("🔬 Model Validation")
    st.caption("Compare FAST / STANDARD / PRECISE voltage model accuracy against a flight log.")

    log_val = st.session_state.get("flight_log")
    if log_val is None:
        st.info("Load a flight log in the **Log Tools** page first (upload or generate synthetic).")
    else:
        st.success(f"Using log: `{Path(log_val.source_file).name}`  "
                   f"({log_val.total_flight_s:.0f} s, {len(log_val.time_s)} samples)")

        pack_ids_v = sorted(db.packs.keys())
        mission_ids_v = sorted(db.missions.keys())
        uav_ids_v = sorted(db.uav_configs.keys())

        if not (pack_ids_v and mission_ids_v and uav_ids_v):
            st.warning("Need at least one pack, mission, and UAV config to run validation.")
        else:
            col_v1, col_v2, col_v3 = st.columns(3)
            with col_v1:
                val_pack_id    = st.selectbox("Battery pack", pack_ids_v, key="val_pack")
                val_mission_id = st.selectbox("Mission", mission_ids_v, key="val_mission")
                val_uav_id     = st.selectbox("UAV config", uav_ids_v, key="val_uav")
            with col_v2:
                val_ambient    = st.number_input("Ambient temp (°C)", -20, 55, 25, key="val_temp")
                val_initial_soc= st.slider("Initial SoC (%)", 80, 100, 100, key="val_soc")
                val_peukert    = st.slider("Peukert k", 1.00, 1.15, 1.05, 0.01, key="val_peukert")
            with col_v3:
                val_cutoff_soc = st.slider("Cutoff SoC (%)", 0, 20, 10, key="val_cutoff")
                val_modes = st.multiselect(
                    "Modes to compare",
                    ["FAST", "STANDARD", "PRECISE"],
                    default=["FAST", "STANDARD"],
                    key="val_modes",
                )

            # Check if PRECISE ECM params available
            from batteries.ecm_store import LogRegistry, get_ecm_for_temperature
            _val_registry = LogRegistry()
            _val_registry.load()
            fitted_packs = list(set(e.pack_id for e in _val_registry.all_entries()))
            if "PRECISE" in val_modes and val_pack_id not in fitted_packs:
                st.warning(
                    f"No fitted ECM params for `{val_pack_id}`. "
                    "Use the **Log Tools** page to fit and register a log first, "
                    "or deselect PRECISE mode."
                )

            if st.button("▶ Run Validation", type="primary", key="val_run"):
                from batteries.voltage_model import ModelMode
                from batteries.model_validator import validate_against_log, plot_validation

                val_pack    = db.packs[val_pack_id]
                val_mission = db.missions[val_mission_id]
                val_uav     = db.uav_configs[val_uav_id]

                mode_map = {
                    "FAST":     ModelMode.FAST,
                    "STANDARD": ModelMode.STANDARD,
                    "PRECISE":  ModelMode.PRECISE,
                }
                val_results = {}
                progress_v = st.progress(0)
                for idx, mode_name in enumerate(val_modes):
                    progress_v.progress(idx / len(val_modes),
                                        text=f"Validating {mode_name}...")
                    ecm_p = get_ecm_for_temperature(_val_registry, val_pack_id, val_ambient) if mode_name == "PRECISE" else None
                    try:
                        val_results[mode_name] = validate_against_log(
                            pack=val_pack, log=log_val,
                            mission=val_mission, uav=val_uav,
                            discharge_pts=db.discharge_pts,
                            mode=mode_map[mode_name],
                            ecm_params=ecm_p,
                            ambient_temp_c=val_ambient,
                            initial_soc_pct=float(val_initial_soc),
                            peukert_k=val_peukert,
                            cutoff_soc_pct=float(val_cutoff_soc),
                        )
                    except Exception as ve:
                        st.error(f"{mode_name} failed: {ve}")
                progress_v.progress(1.0, text="Done!")
                st.session_state["val_results"] = val_results

            val_results = st.session_state.get("val_results", {})
            if val_results:
                # Metrics table
                st.subheader("Accuracy Metrics")
                metric_rows = []
                for mname, vr in val_results.items():
                    metric_rows.append({
                        "Mode":    mname,
                        "RMSE (V)": round(vr.rmse_v, 5),
                        "MAE (V)":  round(vr.mae_v, 5),
                        "R²":       round(vr.r_squared, 5),
                        "Bias (V)": round(vr.bias_v, 5),
                        "Points":   vr.n_points,
                    })
                st.dataframe(pd.DataFrame(metric_rows), use_container_width=True, hide_index=True)

                # Plot
                st.subheader("Voltage Comparison")
                from batteries.model_validator import plot_validation
                fig_val = plot_validation(
                    val_results,
                    title=f"Model Validation — {val_pack_id}",
                )
                st.pyplot(fig_val, use_container_width=True)
                plt.close(fig_val)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 — BULK DATA UPLOAD
# ══════════════════════════════════════════════════════════════════════════════
with tab_bulk:
    st.subheader("📤 Bulk Data Upload")
    st.caption("Import battery data via CSV bulk upload, web scraper, or PDF datasheet.")

    bulk_sub_csv, bulk_sub_web, bulk_sub_pdf = st.tabs(["CSV Bulk Import", "Web Scraper", "PDF Datasheet"])

with bulk_sub_csv:
    st.subheader("CSV Bulk Import")
    st.caption("Download the template, fill it in, then upload to import multiple packs at once.")

    template_cols = [
        "battery_id", "name", "cell_id", "chemistry_id",
        "cells_series", "cells_parallel",
        "pack_voltage_nom", "pack_voltage_max", "pack_voltage_cutoff",
        "pack_capacity_ah", "pack_energy_wh", "pack_weight_g",
        "max_cont_discharge_a", "internal_resistance_mohm",
        "cycle_life", "uav_class", "notes",
    ]
    example_row = {
        "battery_id": "CUSTOM_6S_10AH_V1",
        "name": "Custom 6S 10Ah Pack",
        "cell_id": "CELL_SAMSUNG_50E",
        "chemistry_id": "Li-Ion",
        "cells_series": 6,
        "cells_parallel": 5,
        "pack_voltage_nom": 21.6,
        "pack_voltage_max": 25.2,
        "pack_voltage_cutoff": 18.0,
        "pack_capacity_ah": 10.0,
        "pack_energy_wh": 216.0,
        "pack_weight_g": 1050,
        "max_cont_discharge_a": 50,
        "internal_resistance_mohm": 35.0,
        "cycle_life": 500,
        "uav_class": "Heavy VTOL",
        "notes": "Custom build example",
    }
    template_df = pd.DataFrame([example_row], columns=template_cols)
    template_csv = template_df.to_csv(index=False).encode("utf-8")

    st.download_button(
        "⬇ Download CSV Template",
        data=template_csv,
        file_name="battery_upload_template.csv",
        mime="text/csv",
        key="bulk_dl_template",
    )

    st.divider()

    uploaded_csv = st.file_uploader(
        "Upload filled CSV",
        type=["csv"],
        help="Use the template above. The battery_id column must be unique.",
        key="bulk_upload",
    )

    if uploaded_csv is not None:
        try:
            upload_df = pd.read_csv(uploaded_csv)
            st.write(f"**Preview — {len(upload_df)} rows**")
            st.dataframe(upload_df.head(20), use_container_width=True, hide_index=True)

            missing_cols = [c for c in ["battery_id", "pack_energy_wh", "pack_weight_g"]
                            if c not in upload_df.columns]
            if missing_cols:
                st.error(f"Missing required columns: {', '.join(missing_cols)}")
            else:
                duplicates = [bid for bid in upload_df["battery_id"].astype(str) if bid in db.packs]
                if duplicates:
                    st.warning(f"These IDs already exist and will be skipped: "
                               f"{', '.join(duplicates[:10])}"
                               + ("…" if len(duplicates) > 10 else ""))

                rows_to_add = upload_df[~upload_df["battery_id"].astype(str).isin(set(db.packs.keys()))]
                st.write(f"**{len(rows_to_add)}** new packs will be added, "
                         f"**{len(duplicates)}** will be skipped.")

                if len(rows_to_add) > 0:
                    if st.button(f"💾 Import {len(rows_to_add)} packs",
                                 type="primary", key="bulk_import"):
                        from openpyxl import load_workbook
                        from ui.config import DB_PATH

                        try:
                            wb = load_workbook(DB_PATH)
                            ws = wb["Battery_Catalog"]
                            added = 0
                            errors = []
                            for _, row in rows_to_add.iterrows():
                                def _g(col, default=None):
                                    return row[col] if col in row.index and pd.notna(row[col]) else default

                                try:
                                    s       = int(_g("cells_series", 1))
                                    p       = int(_g("cells_parallel", 1))
                                    nom_v   = float(_g("pack_voltage_nom", 0))
                                    cap_ah  = float(_g("pack_capacity_ah", 0))
                                    en_wh   = float(_g("pack_energy_wh", 0))
                                    wt_g    = float(_g("pack_weight_g", 0))
                                    sp_e    = round(en_wh / (wt_g / 1000), 1) if wt_g > 0 else 0
                                    max_i   = float(_g("max_cont_discharge_a", 0))
                                    max_w   = round(max_i * nom_v, 0) if nom_v > 0 else 0
                                    c_rate  = round(max_i / cap_ah, 2) if cap_ah > 0 else 0
                                    ws.append([
                                        str(_g("battery_id", "")),
                                        str(_g("name", "")),
                                        str(_g("cell_id", "")),
                                        str(_g("chemistry_id", "")),
                                        s, p, s * p,
                                        nom_v,
                                        float(_g("pack_voltage_max", 0)),
                                        float(_g("pack_voltage_cutoff", 0)),
                                        cap_ah, en_wh, wt_g,
                                        sp_e, 0, 0,
                                        max_i, max_w, c_rate,
                                        float(_g("internal_resistance_mohm", 0)),
                                        int(_g("cycle_life", 0)),
                                        str(_g("uav_class", "")),
                                        str(_g("notes", "")),
                                    ])
                                    added += 1
                                except Exception as row_err:
                                    errors.append(f"{row.get('battery_id', '?')}: {row_err}")

                            wb.save(DB_PATH)
                            reload_db()

                            if added:
                                st.success(f"Successfully imported {added} battery pack(s)!")
                            if errors:
                                st.warning(f"Errors on {len(errors)} row(s):\n" +
                                           "\n".join(errors[:10]))
                            st.rerun()
                        except Exception as e:
                            st.error(f"Import failed: {e}")
        except Exception as e:
            st.error(f"Failed to read CSV: {e}")

# ── Web Scraper sub-tab ───────────────────────────────────────────────────────
with bulk_sub_web:
    st.subheader("🌐 Web Scraper")
    st.caption("Point the scraper at a product/spec page to extract battery data automatically. "
               "Review and edit the extracted fields before saving.")

    scrape_url = st.text_input("Target URL", placeholder="https://example.com/battery-spec-page",
                               key="scrape_url")
    col_sc1, col_sc2 = st.columns([1, 3])
    with col_sc1:
        scrape_btn = st.button("🔍 Scrape Page", type="primary", key="scrape_btn",
                               disabled=not scrape_url)

    if scrape_btn and scrape_url:
        try:
            import requests
            from bs4 import BeautifulSoup
            import re

            with st.spinner(f"Fetching {scrape_url}…"):
                headers = {"User-Agent": "Mozilla/5.0 (compatible; BattSim/1.0)"}
                resp = requests.get(scrape_url, headers=headers, timeout=15)
                resp.raise_for_status()
                soup = BeautifulSoup(resp.text, "html.parser")

            # Extract all text tables as DataFrames
            tables = soup.find_all("table")
            extracted_tables = []
            for tbl in tables:
                try:
                    df_list = pd.read_html(str(tbl))
                    if df_list:
                        extracted_tables.append(df_list[0])
                except Exception:
                    pass

            # Extract key-value pairs from spec lists / definition lists / paragraphs
            raw_text = soup.get_text(separator="\n")
            kv_pairs: dict[str, str] = {}

            # Common patterns: "Capacity: 10000mAh", "Weight: 1050g", etc.
            patterns = [
                (r"[Cc]apacit[y|ies][\s:]+([0-9]+(?:\.[0-9]+)?)\s*(mAh|Ah)",   "capacity"),
                (r"[Vv]oltage[\s:]+([0-9]+(?:\.[0-9]+)?)\s*V",                  "voltage_nom"),
                (r"[Ww]eight[\s:]+([0-9]+(?:\.[0-9]+)?)\s*g",                   "weight_g"),
                (r"[Ee]nergy[\s:]+([0-9]+(?:\.[0-9]+)?)\s*Wh",                 "energy_wh"),
                (r"[Cc]ell[s]?\s+[Cc]onfiguration[\s:]+([0-9]+S[0-9]+P?)",     "cell_config"),
                (r"[Mm]ax\s+[Dd]ischarge[\s:]+([0-9]+(?:\.[0-9]+)?)\s*A",      "max_discharge_a"),
                (r"[Cc]hemistr[y|ies][\s:]+(\w[\w\-]+)",                        "chemistry"),
            ]
            for pat, key in patterns:
                m = re.search(pat, raw_text)
                if m:
                    kv_pairs[key] = m.group(1)

            st.session_state["scrape_result_tables"] = extracted_tables
            st.session_state["scrape_result_kv"]     = kv_pairs
            st.session_state["scrape_result_url"]    = scrape_url
            st.success(f"Scraped {len(extracted_tables)} table(s). Review below.")

        except ImportError:
            st.error("Missing dependency: `pip install requests beautifulsoup4`")
        except Exception as e:
            st.error(f"Scrape failed: {e}")

    if st.session_state.get("scrape_result_url"):
        st.caption(f"Results from: {st.session_state['scrape_result_url']}")

        extracted_tables = st.session_state.get("scrape_result_tables", [])
        kv_pairs = st.session_state.get("scrape_result_kv", {})

        if kv_pairs:
            st.write("**Auto-extracted specifications:**")
            kv_df = pd.DataFrame([{"Field": k, "Value": v} for k, v in kv_pairs.items()])
            st.dataframe(kv_df, use_container_width=True, hide_index=True)

        if extracted_tables:
            st.write(f"**Tables found on page ({len(extracted_tables)}):**")
            for i, tdf in enumerate(extracted_tables[:5]):
                with st.expander(f"Table {i+1} ({tdf.shape[0]} rows × {tdf.shape[1]} cols)",
                                 expanded=(i == 0)):
                    st.dataframe(tdf, use_container_width=True, hide_index=True)

        st.divider()
        st.write("**Map extracted data to battery fields and save:**")
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            w_bid   = st.text_input("Battery ID*", key="web_bid")
            w_name  = st.text_input("Name", key="web_name")
            w_chem  = st.text_input("Chemistry", value=kv_pairs.get("chemistry", ""), key="web_chem")
            w_cid   = st.text_input("Cell ID", key="web_cell_id")
        with col_f2:
            _cap_raw = kv_pairs.get("capacity", "0")
            _cap_f   = float(_cap_raw) if _cap_raw.replace(".", "").isdigit() else 0.0
            _volt_f  = float(kv_pairs.get("voltage_nom", "0") or 0)
            _wt_f    = float(kv_pairs.get("weight_g", "0") or 0)
            w_s      = st.number_input("Cells series (S)", 1, 28, 6, key="web_s")
            w_p      = st.number_input("Cells parallel (P)", 1, 50, 1, key="web_p")
            w_cap_ah = st.number_input("Capacity (Ah)", 0.0, 500.0, _cap_f / 1000 if _cap_f > 100 else _cap_f,
                                       0.1, key="web_cap")
        with col_f3:
            w_nom_v  = st.number_input("Nom voltage (V)", 0.0, 150.0, _volt_f, 0.1, key="web_nom_v")
            w_wt_g   = st.number_input("Weight (g)", 0.0, 50000.0, _wt_f, 1.0, key="web_wt_g")
            w_en_wh  = st.number_input("Energy (Wh)", 0.0, 100000.0,
                                       float(kv_pairs.get("energy_wh", 0) or 0), 0.1, key="web_en_wh")
            w_max_i  = st.number_input("Max discharge (A)", 0.0, 1000.0,
                                       float(kv_pairs.get("max_discharge_a", 0) or 0), 0.5, key="web_max_i")

        if st.button("💾 Save Scraped Pack to Database", type="primary", key="web_save_btn"):
            if not w_bid:
                st.error("Battery ID is required.")
            elif w_bid in db.packs:
                st.error(f"Battery ID `{w_bid}` already exists.")
            else:
                try:
                    from openpyxl import load_workbook
                    from ui.config import DB_PATH as _DB
                    sp_e   = round(w_en_wh / (w_wt_g / 1000), 1) if w_wt_g > 0 else 0
                    max_w  = round(w_max_i * w_nom_v, 0)
                    c_rate = round(w_max_i / w_cap_ah, 2) if w_cap_ah > 0 else 0
                    wb = load_workbook(_DB)
                    ws = wb["Battery_Catalog"]
                    ws.append([
                        w_bid, w_name, w_cid, w_chem,
                        w_s, w_p, w_s * w_p,
                        w_nom_v, 0, 0,
                        w_cap_ah, w_en_wh, w_wt_g,
                        sp_e, 0, 0,
                        w_max_i, max_w, c_rate, 0, 0,
                        "", f"Scraped from {st.session_state.get('scrape_result_url', '')}",
                    ])
                    wb.save(_DB)
                    reload_db()
                    st.success(f"Pack `{w_bid}` saved to database!")
                    st.session_state.pop("scrape_result_tables", None)
                    st.session_state.pop("scrape_result_kv", None)
                    st.session_state.pop("scrape_result_url", None)
                    st.rerun()
                except Exception as e:
                    st.error(f"Failed to save: {e}")

# ── PDF Datasheet sub-tab ─────────────────────────────────────────────────────
with bulk_sub_pdf:
    st.subheader("📄 PDF Datasheet Upload")
    st.caption("Upload a battery datasheet PDF. The tool will extract text and attempt to parse "
               "key specifications. Review and edit before saving.")

    pdf_file = st.file_uploader("Upload PDF datasheet", type=["pdf"], key="pdf_upload")

    if pdf_file is not None:
        try:
            import pdfplumber
            import re

            with st.spinner("Extracting text from PDF…"):
                pdf_bytes = pdf_file.read()
                buf = io.BytesIO(pdf_bytes)
                full_text = ""
                pdf_tables = []
                with pdfplumber.open(buf) as pdf:
                    for page in pdf.pages:
                        full_text += page.extract_text() or ""
                        tbls = page.extract_tables()
                        for t in (tbls or []):
                            if t:
                                pdf_tables.append(pd.DataFrame(t[1:], columns=t[0]
                                                               if t[0] else None))

            st.success(f"Extracted {len(full_text)} characters, {len(pdf_tables)} table(s).")

            # Show raw text preview
            with st.expander("📄 Extracted Text (first 3000 chars)", expanded=False):
                st.text(full_text[:3000])

            # Show tables
            if pdf_tables:
                st.write(f"**Tables extracted ({len(pdf_tables)}):**")
                for i, tdf in enumerate(pdf_tables[:5]):
                    with st.expander(f"Table {i+1}", expanded=(i == 0)):
                        st.dataframe(tdf, use_container_width=True, hide_index=True)

            # Auto-parse specs from text
            patterns = [
                (r"[Cc]apacit[y][\s:]+([0-9]+(?:\.[0-9]+)?)\s*(mAh|Ah)",  "capacity"),
                (r"[Nn]ominal\s+[Vv]oltage[\s:]+([0-9]+(?:\.[0-9]+)?)\s*V", "voltage_nom"),
                (r"[Ww]eight[\s:]+([0-9]+(?:\.[0-9]+)?)\s*g",               "weight_g"),
                (r"[Ee]nergy[\s:]+([0-9]+(?:\.[0-9]+)?)\s*Wh",              "energy_wh"),
                (r"([0-9]+S[0-9]+P?)",                                       "cell_config"),
                (r"[Mm]ax\s+[Dd]ischarge[\s:]+([0-9]+(?:\.[0-9]+)?)\s*A",  "max_discharge_a"),
            ]
            kv: dict[str, str] = {}
            for pat, key in patterns:
                m = re.search(pat, full_text)
                if m:
                    kv[key] = m.group(1)

            if kv:
                st.write("**Auto-extracted specifications:**")
                st.dataframe(pd.DataFrame([{"Field": k, "Value": v} for k, v in kv.items()]),
                             use_container_width=True, hide_index=True)

            st.divider()
            st.write("**Review and save:**")
            col_p1, col_p2, col_p3 = st.columns(3)
            with col_p1:
                p_bid  = st.text_input("Battery ID*", key="pdf_bid",
                                       value=Path(pdf_file.name).stem[:40].replace(" ", "_").upper())
                p_name = st.text_input("Name", value=Path(pdf_file.name).stem, key="pdf_name")
                p_chem = st.text_input("Chemistry", key="pdf_chem")
            with col_p2:
                _cap   = float(kv.get("capacity", "0") or 0)
                p_s    = st.number_input("Cells series (S)", 1, 28, 6, key="pdf_s")
                p_p    = st.number_input("Cells parallel (P)", 1, 50, 1, key="pdf_p")
                p_cap  = st.number_input("Capacity (Ah)", 0.0, 500.0,
                                         _cap / 1000 if _cap > 100 else _cap, 0.1, key="pdf_cap")
            with col_p3:
                p_v   = st.number_input("Nom voltage (V)", 0.0, 150.0,
                                        float(kv.get("voltage_nom", 0) or 0), 0.1, key="pdf_v")
                p_wt  = st.number_input("Weight (g)", 0.0, 50000.0,
                                        float(kv.get("weight_g", 0) or 0), 1.0, key="pdf_wt")
                p_en  = st.number_input("Energy (Wh)", 0.0, 100000.0,
                                        float(kv.get("energy_wh", 0) or 0), 0.1, key="pdf_en")
                p_mi  = st.number_input("Max discharge (A)", 0.0, 1000.0,
                                        float(kv.get("max_discharge_a", 0) or 0), 0.5, key="pdf_mi")

            if st.button("💾 Save PDF Pack to Database", type="primary", key="pdf_save_btn"):
                if not p_bid:
                    st.error("Battery ID is required.")
                elif p_bid in db.packs:
                    st.error(f"Battery ID `{p_bid}` already exists.")
                else:
                    try:
                        from openpyxl import load_workbook
                        from ui.config import DB_PATH as _DB2
                        sp_e   = round(p_en / (p_wt / 1000), 1) if p_wt > 0 else 0
                        max_w  = round(p_mi * p_v, 0)
                        c_rate = round(p_mi / p_cap, 2) if p_cap > 0 else 0
                        wb = load_workbook(_DB2)
                        ws = wb["Battery_Catalog"]
                        ws.append([
                            p_bid, p_name, "", p_chem,
                            p_s, p_p, p_s * p_p,
                            p_v, 0, 0,
                            p_cap, p_en, p_wt,
                            sp_e, 0, 0,
                            p_mi, max_w, c_rate, 0, 0,
                            "", f"From PDF: {pdf_file.name}",
                        ])
                        wb.save(_DB2)
                        reload_db()
                        st.success(f"Pack `{p_bid}` saved to database!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Failed to save: {e}")

        except ImportError:
            st.error("Missing dependency: `pip install pdfplumber`")
        except Exception as e:
            st.error(f"Failed to parse PDF: {e}")
    else:
        st.info("Upload a PDF datasheet to begin extraction.")

