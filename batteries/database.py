"""
batteries/database.py
Read and write battery_db.xlsx — loads all entities into Python dataclasses.
"""
from __future__ import annotations
from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from batteries.models import (
    Chemistry, Cell, BatteryPack, DischargePoint,
    Equipment, UAVConfiguration, MissionPhase, MissionProfile,
    EquipmentPhaseAssignment,
)


def _safe(val, typ=float, default=None):
    """Safely cast a cell value; return default on failure."""
    if val is None or val == "N/A" or val == "":
        return default
    try:
        return typ(val)
    except (TypeError, ValueError):
        return default


def _str(val, default="") -> str:
    return str(val).strip() if val is not None else default


class BatteryDatabase:
    def __init__(self, path: str | Path = "battery_db.xlsx"):
        self.path = Path(path)
        self._wb = None
        # In-memory caches
        self.chemistries:   dict[str, Chemistry]      = {}
        self.cells:         dict[str, Cell]           = {}
        self.packs:         dict[str, BatteryPack]    = {}
        self.discharge_pts: list[DischargePoint]      = []
        self.equipment:     dict[str, Equipment]      = {}
        self.uav_configs:   dict[str, UAVConfiguration] = {}
        self.missions:      dict[str, MissionProfile] = {}

    # ── Load / reload ──────────────────────────────────────────────────────

    def load(self) -> "BatteryDatabase":
        """Load all sheets from the workbook into memory."""
        self._wb = load_workbook(self.path, data_only=True)
        self._load_chemistries()
        self._load_cells()
        self._load_packs()
        self._load_discharge_profiles()
        self._load_equipment()
        self._load_uav_configs()
        self._load_missions()
        return self

    def reload(self) -> "BatteryDatabase":
        return self.load()

    # ── Private loaders ────────────────────────────────────────────────────

    def _rows(self, sheet_name: str, start_row: int):
        """Yield non-empty rows from a sheet starting at start_row."""
        ws = self._wb[sheet_name]
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            if all(v is None for v in row):
                continue
            yield row

    def _load_chemistries(self):
        for r in self._rows("Chemistry_Library", 4):
            if not r[0]:
                continue
            self.chemistries[_str(r[0])] = Chemistry(
                chem_id=_str(r[0]),
                name=_str(r[1]),
                short_code=_str(r[2]),
                voltage_nominal=_safe(r[3], float, 0),
                voltage_cutoff=_safe(r[4], float, 0),
                energy_density_wh_kg=_safe(r[5], float, 0),
                energy_density_wh_l=_safe(r[6], float, 0),
                specific_power_w_kg=_safe(r[7], float, 0),
                cycle_life=_safe(r[8], int, 0),
                temp_min_c=_safe(r[9], float, -20),
                temp_max_c=_safe(r[10], float, 60),
                self_discharge_pct_month=_safe(r[11], float, 0),
                max_cont_discharge_c=_safe(r[12], float, 0),
                max_pulse_discharge_c=_safe(r[13], float, 0),
                charge_efficiency_pct=_safe(r[14], float, 99),
                safety_rating=_str(r[15]),
                relative_cost=_str(r[16]),
                notes=_str(r[17]),
            )

    def _load_cells(self):
        for r in self._rows("Cell_Catalog", 3):
            if not r[0]:
                continue
            self.cells[_str(r[0])] = Cell(
                cell_id=_str(r[0]),
                manufacturer=_str(r[1]),
                model=_str(r[2]),
                chemistry_id=_str(r[3]),
                cell_format=_str(r[4]),
                voltage_nominal=_safe(r[5], float, 0),
                voltage_max=_safe(r[6], float, 0),
                voltage_cutoff=_safe(r[7], float, 0),
                capacity_ah=_safe(r[8], float, 0),
                energy_wh=_safe(r[9], float, 0),
                weight_g=_safe(r[10], float, 0),
                specific_energy_wh_kg=_safe(r[11], float, 0),
                volume_cm3=_safe(r[12], float),
                energy_density_wh_l=_safe(r[13], float),
                max_cont_discharge_a=_safe(r[14], float, 0),
                max_pulse_discharge_a=_safe(r[15], float, 0),
                max_charge_rate_c=_safe(r[16], float, 0),
                internal_resistance_mohm=_safe(r[17], float, 0),
                cycle_life=_safe(r[18], int, 0),
                notes=_str(r[19]),
            )

    def _load_packs(self):
        for r in self._rows("Battery_Catalog", 3):
            if not r[0]:
                continue
            self.packs[_str(r[0])] = BatteryPack(
                battery_id=_str(r[0]),
                name=_str(r[1]),
                cell_id=_str(r[2]),
                chemistry_id=_str(r[3]),
                cells_series=_safe(r[4], int, 1),
                cells_parallel=_safe(r[5], int, 1),
                pack_voltage_nom=_safe(r[7], float, 0),
                pack_voltage_max=_safe(r[8], float, 0),
                pack_voltage_cutoff=_safe(r[9], float, 0),
                pack_capacity_ah=_safe(r[10], float, 0),
                pack_energy_wh=_safe(r[11], float, 0),
                pack_weight_g=_safe(r[12], float, 0),
                specific_energy_wh_kg=_safe(r[13], float, 0),
                pack_volume_cm3=_safe(r[14], float),
                energy_density_wh_l=_safe(r[15], float),
                max_cont_discharge_a=_safe(r[16], float, 0),
                max_cont_discharge_w=_safe(r[17], float, 0),
                cont_c_rate=_safe(r[18], float, 0),
                internal_resistance_mohm=_safe(r[19], float, 0),
                cycle_life=_safe(r[20], int, 0),
                uav_class=_str(r[21]),
                notes=_str(r[22]),
            )

    def _load_discharge_profiles(self):
        for r in self._rows("Discharge_Profiles", 3):
            if not r[0]:
                continue
            self.discharge_pts.append(DischargePoint(
                chem_id=_str(r[0]),
                c_rate=_safe(r[1], float, 1.0),
                temperature_c=_safe(r[2], float, 25.0),
                soc_pct=_safe(r[3], float, 0),
                voltage_v=_safe(r[4], float, 0),
                normalised_capacity_pct=_safe(r[5], float, 0),
            ))

    def _load_equipment(self):
        import warnings

        # Detect legacy column format by inspecting the header row (row 2).
        # Old format has hover_power_w in col 8 (1-based); new has operating_power_w.
        ws_eq = self._wb["Equipment_DB"]
        header_col8 = _str(ws_eq.cell(row=2, column=8).value).lower()
        legacy_format = any(kw in header_col8 for kw in ("hover", "climb", "cruise", "legacy"))

        if legacy_format:
            warnings.warn(
                "Equipment_DB sheet has old per-phase power columns "
                "(hover/climb/cruise). Auto-migrating to 3-level model on load. "
                "Save equipment from the Equipment Editor to write the new format.",
                UserWarning, stacklevel=3,
            )

        for r in self._rows("Equipment_DB", 3):
            if not r[0]:
                continue
            if legacy_format:
                # Old layout: col6=idle, col7=hover, col8=climb, col9=cruise,
                #             col10=max, col11=weight, col12=eff, col13=duty,
                #             col14=notes, col15=active
                idle_w   = _safe(r[6], float, 0)
                hover_w  = _safe(r[7], float, 0)
                climb_w  = _safe(r[8], float, 0)
                cruise_w = _safe(r[9], float, 0)
                max_w    = _safe(r[10], float, 0)
                # Derive operating_power_w as the max of the three phase powers
                operating_w = max(hover_w, climb_w, cruise_w)
                weight    = _safe(r[11], float, 0)
                eff       = _safe(r[12], float, 100)
                duty      = _safe(r[13], float, 100)
                notes_v   = _str(r[14])
                active_v  = _str(r[15]).upper() == "YES"
            else:
                # New layout: col6=idle, col7=operating, col8=max,
                #             col9=weight, col10=eff, col11=duty,
                #             col12=notes, col13=active
                idle_w      = _safe(r[6], float, 0)
                operating_w = _safe(r[7], float, 0)
                max_w       = _safe(r[8], float, 0)
                weight      = _safe(r[9], float, 0)
                eff         = _safe(r[10], float, 100)
                duty        = _safe(r[11], float, 100)
                notes_v     = _str(r[12])
                active_v    = _str(r[13]).upper() == "YES"

            self.equipment[_str(r[0])] = Equipment(
                equip_id=_str(r[0]),
                category=_str(r[1]),
                manufacturer=_str(r[2]),
                model=_str(r[3]),
                nom_voltage_v=_safe(r[4], float, 0),
                nom_current_a=_safe(r[5], float, 0),
                idle_power_w=idle_w,
                operating_power_w=operating_w,
                max_power_w=max_w,
                weight_g=weight,
                efficiency_pct=eff,
                duty_cycle_pct=duty,
                notes=notes_v,
                active=active_v,
            )

    def _load_uav_configs(self):
        """Build UAVConfiguration objects from UAV_Configurations sheet."""
        configs: dict[str, dict] = {}  # uav_id → {name, items}
        for r in self._rows("UAV_Configurations", 4):
            if not r[0]:
                continue
            uid = _str(r[0])
            if uid not in configs:
                configs[uid] = {"name": _str(r[1]), "items": []}
            equip_id = _str(r[2])
            qty = _safe(r[3], int, 1)
            eq = self.equipment.get(equip_id)
            if eq:
                configs[uid]["items"].append((eq, qty))

        for uid, data in configs.items():
            cfg = UAVConfiguration(
                uav_id=uid,
                name=data["name"],
                equipment_list=data["items"],
            )
            self.uav_configs[uid] = cfg

    def _validate_phase_type(self, phase_type: str, mission_id: str, phase_name: str) -> str:
        """Return phase_type upper-cased; warn and default to CRUISE if unrecognised."""
        from batteries.ardupilot_modes import ALL_VALID_PHASE_TYPES
        import warnings
        upper = phase_type.upper()
        if upper not in ALL_VALID_PHASE_TYPES:
            warnings.warn(
                f"Mission '{mission_id}' phase '{phase_name}': "
                f"unknown phase_type '{phase_type}' — defaulting to 'CRUISE'.",
                UserWarning, stacklevel=4,
            )
            return "CRUISE"
        return upper

    def _load_missions(self):
        """Build MissionProfile objects from Mission_Profiles sheet."""
        phases_by_mission: dict[str, list[MissionPhase]] = {}
        meta: dict[str, tuple[str, str]] = {}  # mission_id → (name, uav_id)

        for r in self._rows("Mission_Profiles", 4):
            if not r[0] or not r[4]:
                continue
            mid = _str(r[0])
            if mid not in meta:
                meta[mid] = (_str(r[1]), _str(r[2]))
            if mid not in phases_by_mission:
                phases_by_mission[mid] = []

            phases_by_mission[mid].append(MissionPhase(
                mission_id=mid,
                mission_name=_str(r[1]),
                uav_config_id=_str(r[2]),
                phase_seq=_safe(r[3], int, 0),
                phase_name=_str(r[4]),
                phase_type=self._validate_phase_type(_str(r[5]), mid, _str(r[4])),
                duration_s=_safe(r[6], float, 0),
                distance_m=_safe(r[7], float, 0),
                altitude_m=_safe(r[8], float, 0),
                airspeed_ms=_safe(r[9], float, 0),
                power_override_w=_safe(r[10], float),
                notes=_str(r[11]),
            ))

        # Load equipment phase assignments (new sheet; may not exist yet)
        assignments_by_mission: dict[str, list[EquipmentPhaseAssignment]] = {}
        if "Equipment_Phase_Assignments" in self._wb.sheetnames:
            for r in self._rows("Equipment_Phase_Assignments", 2):
                if not r[0]:
                    continue
                mid = _str(r[0])
                if mid not in assignments_by_mission:
                    assignments_by_mission[mid] = []
                assignments_by_mission[mid].append(EquipmentPhaseAssignment(
                    mission_id=mid,
                    phase_seq=_safe(r[1], int, 0),
                    equipment_id=_str(r[2]),
                    state=_str(r[3]) or "on",
                    custom_power_w=_safe(r[4], float),
                    custom_power_pct=_safe(r[5], float),
                ))

        for mid, phases in phases_by_mission.items():
            name, uav_id = meta[mid]
            phases_sorted = sorted(phases, key=lambda p: p.phase_seq)

            # Use loaded assignments, or generate defaults (all equipment "on")
            if mid in assignments_by_mission:
                assignments = assignments_by_mission[mid]
            else:
                assignments = []
                uav_cfg = self.uav_configs.get(uav_id)
                if uav_cfg:
                    for phase in phases_sorted:
                        for eq, _qty in uav_cfg.equipment_list:
                            assignments.append(EquipmentPhaseAssignment(
                                mission_id=mid,
                                phase_seq=phase.phase_seq,
                                equipment_id=eq.equip_id,
                                state="on",
                            ))

            self.missions[mid] = MissionProfile(
                mission_id=mid,
                mission_name=name,
                uav_config_id=uav_id,
                phases=phases_sorted,
                equipment_assignments=assignments,
            )

    # ── Query helpers ──────────────────────────────────────────────────────

    def cells_by_chemistry(self, chem_id: str) -> list[Cell]:
        return [c for c in self.cells.values() if c.chemistry_id == chem_id]

    def packs_by_chemistry(self, chem_id: str) -> list[BatteryPack]:
        return [p for p in self.packs.values() if p.chemistry_id == chem_id]

    def packs_by_uav_class(self, uav_class: str) -> list[BatteryPack]:
        return [p for p in self.packs.values()
                if uav_class.lower() in p.uav_class.lower()]

    def discharge_curve(self, chem_id: str, c_rate: float,
                        temp_c: float = 25.0) -> list[DischargePoint]:
        """Return discharge curve points for a chemistry/C-rate/temp."""
        pts = [p for p in self.discharge_pts
               if p.chem_id == chem_id
               and abs(p.c_rate - c_rate) < 0.01
               and abs(p.temperature_c - temp_c) < 1.0]
        return sorted(pts, key=lambda p: p.soc_pct, reverse=True)

    def equipment_by_category(self, category: str) -> list[Equipment]:
        return [e for e in self.equipment.values()
                if e.category.lower() == category.lower() and e.active]

    def summary(self) -> str:
        lines = [
            "═══ Battery Database Summary ═══",
            f"  Chemistries       : {len(self.chemistries)}",
            f"  Cells             : {len(self.cells)}",
            f"  Battery packs     : {len(self.packs)}",
            f"  Discharge points  : {len(self.discharge_pts)}",
            f"  Equipment items   : {len(self.equipment)}",
            f"  UAV configurations: {len(self.uav_configs)}",
            f"  Mission profiles  : {len(self.missions)}",
        ]
        return "\n".join(lines)

    # ── Write helpers ──────────────────────────────────────────────────────

    def append_custom_pack(self, pack: BatteryPack):
        """Append a custom pack to Battery_Catalog and save."""
        wb = load_workbook(self.path)
        ws = wb["Battery_Catalog"]
        next_row = ws.max_row + 1
        row_vals = [
            pack.battery_id, pack.name, pack.cell_id, pack.chemistry_id,
            pack.cells_series, pack.cells_parallel, pack.total_cells,
            pack.pack_voltage_nom, pack.pack_voltage_max, pack.pack_voltage_cutoff,
            pack.pack_capacity_ah, pack.pack_energy_wh, pack.pack_weight_g,
            pack.specific_energy_wh_kg, pack.pack_volume_cm3, pack.energy_density_wh_l,
            pack.max_cont_discharge_a, pack.max_cont_discharge_w, pack.cont_c_rate,
            pack.internal_resistance_mohm, pack.cycle_life, pack.uav_class, pack.notes,
        ]
        ws.append(row_vals)
        wb.save(self.path)
        self.packs[pack.battery_id] = pack
        print(f"[OK]Pack '{pack.battery_id}' appended to Battery_Catalog.")

    def save_equipment(self, equipment_list: list[Equipment] | None = None) -> None:
        """
        Write equipment to Equipment_DB sheet using the new 3-level column layout.
        If equipment_list is None, uses self.equipment.values().
        Clears all data rows (row 3+) and rewrites from scratch.
        """
        wb = load_workbook(self.path)
        ws = wb["Equipment_DB"]

        # Write new column headers at row 2
        headers = [
            "equip_id", "category", "manufacturer", "model",
            "nom_voltage_v", "nom_current_a",
            "idle_power_w", "operating_power_w", "max_power_w",
            "weight_g", "efficiency_pct", "duty_cycle_pct",
            "notes", "active",
        ]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=2, column=col_idx, value=header)

        # Clear existing data rows
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        items = equipment_list if equipment_list is not None else list(self.equipment.values())
        for row_idx, eq in enumerate(items, start=3):
            ws.cell(row=row_idx, column=1,  value=eq.equip_id)
            ws.cell(row=row_idx, column=2,  value=eq.category)
            ws.cell(row=row_idx, column=3,  value=eq.manufacturer)
            ws.cell(row=row_idx, column=4,  value=eq.model)
            ws.cell(row=row_idx, column=5,  value=eq.nom_voltage_v)
            ws.cell(row=row_idx, column=6,  value=eq.nom_current_a)
            ws.cell(row=row_idx, column=7,  value=eq.idle_power_w)
            ws.cell(row=row_idx, column=8,  value=eq.operating_power_w)
            ws.cell(row=row_idx, column=9,  value=eq.max_power_w)
            ws.cell(row=row_idx, column=10, value=eq.weight_g)
            ws.cell(row=row_idx, column=11, value=eq.efficiency_pct)
            ws.cell(row=row_idx, column=12, value=eq.duty_cycle_pct)
            ws.cell(row=row_idx, column=13, value=eq.notes)
            ws.cell(row=row_idx, column=14, value="YES" if eq.active else "NO")

        wb.save(self.path)
        # Refresh in-memory cache
        for eq in items:
            self.equipment[eq.equip_id] = eq
        print(f"[OK] Equipment_DB saved ({len(items)} items, new 3-level format).")

    def save_equipment_assignments(
        self,
        mission_id: str,
        assignments: list[EquipmentPhaseAssignment],
    ) -> None:
        """
        Persist equipment phase assignments for one mission to the
        Equipment_Phase_Assignments sheet (created if absent).
        Only rows for mission_id are replaced; other missions are preserved.
        """
        wb = load_workbook(self.path)

        if "Equipment_Phase_Assignments" not in wb.sheetnames:
            ws = wb.create_sheet("Equipment_Phase_Assignments")
            ws.append(["mission_id", "phase_seq", "equipment_id",
                        "state", "custom_power_w", "custom_power_pct"])
        else:
            ws = wb["Equipment_Phase_Assignments"]

        # Collect rows NOT belonging to this mission_id
        kept_rows: list[tuple] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            if _str(row[0]) != mission_id:
                kept_rows.append(row)

        # Clear data rows and rewrite
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        write_row = 2
        for r in kept_rows:
            for col_idx, val in enumerate(r, start=1):
                ws.cell(row=write_row, column=col_idx, value=val)
            write_row += 1

        for a in assignments:
            ws.cell(row=write_row, column=1, value=a.mission_id)
            ws.cell(row=write_row, column=2, value=a.phase_seq)
            ws.cell(row=write_row, column=3, value=a.equipment_id)
            ws.cell(row=write_row, column=4, value=a.state)
            ws.cell(row=write_row, column=5, value=a.custom_power_w)
            ws.cell(row=write_row, column=6, value=a.custom_power_pct)
            write_row += 1

        wb.save(self.path)
        # Refresh in-memory mission assignments
        if mission_id in self.missions:
            self.missions[mission_id].equipment_assignments = assignments
        print(f"[OK] Equipment_Phase_Assignments saved for mission '{mission_id}' "
              f"({len(assignments)} rows).")

    def append_custom_cell(self, cell: Cell):
        """Append a custom cell to Cell_Catalog and save."""
        wb = load_workbook(self.path)
        ws = wb["Cell_Catalog"]
        ws.append([
            cell.cell_id, cell.manufacturer, cell.model, cell.chemistry_id,
            cell.cell_format, cell.voltage_nominal, cell.voltage_max,
            cell.voltage_cutoff, cell.capacity_ah, cell.energy_wh,
            cell.weight_g, cell.specific_energy_wh_kg, cell.volume_cm3,
            cell.energy_density_wh_l, cell.max_cont_discharge_a,
            cell.max_pulse_discharge_a, cell.max_charge_rate_c,
            cell.internal_resistance_mohm, cell.cycle_life, cell.notes,
        ])
        wb.save(self.path)
        self.cells[cell.cell_id] = cell
        print(f"[OK]Cell '{cell.cell_id}' appended to Cell_Catalog.")
