"""
mission/report_generator.py

Generate formatted Excel reports from simulation results and flight log analysis.

Sheets produced:
  1. Cover          — title, run metadata
  2. Mission_Summary — per-phase energy/power breakdown table
  3. Battery_Scorecard — multi-pack comparison table
  4. Temp_Sensitivity — temperature sweep results matrix
  5. Log_vs_Sim      — real log overlaid on simulation predictions
  6. Fitted_Params   — reverse-engineered battery parameters
  7. Charts          — embedded matplotlib charts (saved as images then inserted)
"""
from __future__ import annotations
import io
import math
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np
    HAS_MPL = True
except ImportError:
    HAS_MPL = False

from batteries.models import BatteryPack, MissionProfile
from mission.simulator import SimulationResult


# ── Palette & style helpers (matches battery_db.xlsx conventions) ─────────────

C = {
    "hdr_bg":  "1F3864", "hdr_fg":  "FFFFFF",
    "sub_bg":  "2E75B6", "sub_fg":  "FFFFFF",
    "acc1":    "41719C", "acc2":    "BDD7EE",
    "acc3":    "D6E4F0", "yellow":  "FFF2CC",
    "green":   "E2EFDA", "orange":  "FCE4D6",
    "white":   "FFFFFF", "gray_l":  "F2F2F2",
    "red_l":   "FFCCCC", "blue_in": "0000FF",
    "black":   "000000",
    # Phase colours
    "IDLE":        "EEEEEE", "TAKEOFF":    "FFD9B3",
    "CLIMB":       "FFF0B3", "CRUISE":     "C8EBD4",
    "HOVER":       "C0D9F5", "DESCEND":    "D5E4F7",
    "LAND":        "E8D9F5", "PAYLOAD_OPS":"FFD0DD",
    "EMERGENCY":   "FFB3B3",
}

def _thin():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def _ap(cell, font=None, fill=None, align=None):
    if font:  cell.font      = font
    if fill:  cell.fill      = fill
    if align: cell.alignment = align
    cell.border = _thin()

def _hdr(bg=C["hdr_bg"], fg=C["hdr_fg"], sz=11, bold=True, wrap=True):
    return (Font(name="Arial", bold=bold, color=fg, size=sz),
            PatternFill("solid", fgColor=bg),
            Alignment(horizontal="center", vertical="center", wrap_text=wrap))

def _plain(bg=C["white"], bold=False, center=True, color=C["black"]):
    return (Font(name="Arial", color=color, size=10, bold=bold),
            PatternFill("solid", fgColor=bg),
            Alignment(horizontal="center" if center else "left",
                      vertical="center"))

def _scw(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def _merge_hdr(ws, rng, text, bg=C["hdr_bg"], sz=13):
    ws.merge_cells(rng)
    c = ws[rng.split(":")[0]]
    c.value = text
    _ap(c, *_hdr(bg=bg, sz=sz))
    return c


# ── Chart helpers ─────────────────────────────────────────────────────────────

def _chart_to_image(fig) -> io.BytesIO:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=120, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    return buf


def _insert_image(ws, img_buf: io.BytesIO, anchor: str, width_px=550):
    img = XLImage(img_buf)
    img.width  = width_px
    img.height = int(width_px * 0.55)
    ws.add_image(img, anchor)


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_cover(wb: Workbook, meta: dict):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 60
    ws.row_dimensions[1].height = 70
    ws.row_dimensions[3].height = 30

    _merge_hdr(ws, "A1:C1", "UAV Battery Simulation Report", sz=18)
    for row, (label, val) in enumerate([
        ("Generated",       datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Pack(s)",         meta.get("packs", "")),
        ("Mission",         meta.get("mission", "")),
        ("UAV",             meta.get("uav", "")),
        ("Ambient Temp",    f"{meta.get('temp_c', 25.0)} °C"),
        ("Peukert k",       str(meta.get("peukert_k", 1.05))),
        ("Sheets",          "Mission Summary · Battery Scorecard · Temp Sensitivity · Log vs Sim · Fitted Params"),
    ], start=3):
        ws.row_dimensions[row].height = 22
        c1 = ws.cell(row=row, column=1, value=label)
        _ap(c1, *_plain(bg=C["acc2"], bold=True, center=False))
        c2 = ws.cell(row=row, column=2, value=val)
        c2.font = Font(name="Arial", size=10)
        c2.border = _thin()
        ws.merge_cells(f"B{row}:C{row}")


def _build_mission_summary(
    wb: Workbook,
    result: SimulationResult,
    pack: BatteryPack,
    mission: MissionProfile,
):
    ws = wb.create_sheet("Mission_Summary")
    ws.sheet_view.showGridLines = False

    _merge_hdr(ws, "A1:I1",
               f"Mission Summary — {pack.battery_id} × {mission.mission_id}")
    ws.row_dimensions[1].height = 36

    headers = ["Phase #", "Phase Name", "Phase Type", "Duration (s)",
               "Avg Power (W)", "Energy (Wh)", "Start SoC (%)",
               "End SoC (%)", "Min Voltage (V)"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        _ap(c, *_hdr(bg=C["sub_bg"], sz=9))
    ws.row_dimensions[2].height = 36

    # Group result by phase transitions
    phases_data = []
    if result.phase_type:
        prev = result.phase_type[0]
        start_idx = 0
        for idx, ph in enumerate(result.phase_type):
            if ph != prev or idx == len(result.phase_type) - 1:
                end_idx = idx
                seg_v   = result.voltage_v[start_idx:end_idx]
                seg_soc = result.soc_pct[start_idx:end_idx]
                seg_pw  = result.power_w[start_idx:end_idx]
                dur     = result.time_s[end_idx-1] - result.time_s[start_idx]
                phases_data.append({
                    "type":     prev,
                    "dur_s":    round(dur, 1),
                    "avg_pw":   round(sum(seg_pw)/max(1,len(seg_pw)), 1),
                    "energy":   round(sum(p*result.dt_s/3600 for p in seg_pw), 2),
                    "soc_start":round(seg_soc[0], 1) if seg_soc else 0,
                    "soc_end":  round(seg_soc[-1], 1) if seg_soc else 0,
                    "min_v":    round(min(seg_v), 3) if seg_v else 0,
                })
                prev = ph
                start_idx = idx

    for ri, (mp, pd) in enumerate(
        zip(mission.phases, phases_data), start=3
    ):
        ws.row_dimensions[ri].height = 20
        bg = C.get(pd["type"], C["white"])
        for ci, val in enumerate([
            ri - 2, mp.phase_name, pd["type"], pd["dur_s"],
            pd["avg_pw"], pd["energy"], pd["soc_start"],
            pd["soc_end"], pd["min_v"],
        ], 1):
            c = ws.cell(row=ri, column=ci, value=val)
            _ap(c, *_plain(bg=bg, center=(ci != 2)))
            if ci in (5, 6):
                c.number_format = "0.0"
            if ci == 9:
                c.number_format = "0.000"

    # Totals row
    tot_row = 3 + len(phases_data)
    ws.row_dimensions[tot_row].height = 22
    ws.cell(row=tot_row, column=1, value="TOTAL")
    _ap(ws.cell(row=tot_row, column=1), *_hdr(sz=10))
    ws.merge_cells(f"B{tot_row}:C{tot_row}")
    ws.cell(row=tot_row, column=4, value=result.total_duration_s)
    ws.cell(row=tot_row, column=5, value=round(
        sum(p for p in result.power_w) / max(1, len(result.power_w)), 1))
    ws.cell(row=tot_row, column=6, value=round(result.total_energy_consumed_wh, 2))
    ws.cell(row=tot_row, column=8, value=round(result.final_soc, 1))
    ws.cell(row=tot_row, column=9, value=round(result.min_voltage, 3))
    for ci in range(1, 10):
        c = ws.cell(row=tot_row, column=ci)
        _ap(c, *_plain(bg=C["acc2"], bold=True))

    _scw(ws, {"A":7,"B":24,"C":14,"D":11,"E":12,"F":11,"G":11,"H":11,"I":13})


def _build_battery_scorecard(
    wb: Workbook,
    results: list[SimulationResult],
    packs: list[BatteryPack],
    mission_name: str,
    ambient_temp: float,
):
    ws = wb.create_sheet("Battery_Scorecard")
    ws.sheet_view.showGridLines = False
    _merge_hdr(ws, "A1:N1",
               f"Battery Scorecard — {mission_name} @ {ambient_temp}°C")
    ws.row_dimensions[1].height = 36

    headers = ["Pack ID", "Chemistry", "Config", "Energy (Wh)", "Weight (g)",
               "Wh/kg", "Energy used (Wh)", "Final SoC (%)", "Min V (V)",
               "Peak sag (V)", "Max I (A)", "Max T (°C)",
               "Margin (%)", "Status"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        _ap(c, *_hdr(bg=C["sub_bg"], sz=9))
    ws.row_dimensions[2].height = 36

    for ri, (r, p) in enumerate(zip(results, packs), start=3):
        ws.row_dimensions[ri].height = 22
        margin = (r.final_soc - 10) / 90 * 100 if r.final_soc > 10 else 0
        bg = C["green"] if not r.depleted and margin > 15 else (
             C["yellow"] if not r.depleted else C["red_l"])
        status = "PASS" if not r.depleted and margin > 10 else (
                 "MARGINAL" if not r.depleted else "FAIL")
        row_vals = [
            p.battery_id, p.chemistry_id,
            f"{p.cells_series}S{p.cells_parallel}P",
            p.pack_energy_wh, p.pack_weight_g,
            p.specific_energy_wh_kg,
            round(r.total_energy_consumed_wh, 1),
            round(r.final_soc, 1),
            round(r.min_voltage, 3),
            round(r.peak_sag_v, 3),
            round(r.max_current, 1),
            round(r.max_temp_c, 1),
            round(margin, 1),
            status,
        ]
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            cell_bg = bg if ci != 14 else (
                C["green"] if status == "PASS" else
                C["yellow"] if status == "MARGINAL" else C["red_l"])
            _ap(c, *_plain(bg=cell_bg, bold=(ci == 1 or ci == 14)))

    _scw(ws, {"A":20,"B":10,"C":9,"D":10,"E":10,"F":9,"G":13,
              "H":11,"I":10,"J":10,"K":9,"L":9,"N":9,"O":10})


def _build_temp_sensitivity(
    wb: Workbook,
    temps: list[float],
    sweep_results: list[SimulationResult],
    pack_id: str,
    mission_id: str,
):
    ws = wb.create_sheet("Temp_Sensitivity")
    ws.sheet_view.showGridLines = False
    _merge_hdr(ws, "A1:J1",
               f"Temperature Sensitivity — {pack_id} × {mission_id}")
    ws.row_dimensions[1].height = 36

    headers = ["Ambient (°C)", "Final SoC (%)", "Energy Used (Wh)",
               "Min Voltage (V)", "Peak Sag (V)", "Max Current (A)",
               "Max Cell T (°C)", "T Rise (°C)", "Depleted", "Rating"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        _ap(c, *_hdr(bg=C["sub_bg"], sz=9))
    ws.row_dimensions[2].height = 36

    for ri, (t, r) in enumerate(zip(temps, sweep_results), start=3):
        ws.row_dimensions[ri].height = 20
        t_rise = r.max_temp_c - t
        rating = ("PASS"     if not r.depleted and r.final_soc > 20 else
                  "MARGINAL" if not r.depleted else "FAIL")
        bg = (C["green"] if rating == "PASS" else
              C["yellow"] if rating == "MARGINAL" else C["red_l"])
        row_vals = [
            t, round(r.final_soc, 1), round(r.total_energy_consumed_wh, 1),
            round(r.min_voltage, 3), round(r.peak_sag_v, 3),
            round(r.max_current, 1), round(r.max_temp_c, 1),
            round(t_rise, 1), "YES" if r.depleted else "no", rating,
        ]
        for ci, val in enumerate(row_vals, 1):
            cell_bg = bg if ci == 10 else (
                C["red_l"] if ci == 3 and t < 0 else C["white"] if ci != 1 else
                C["acc2"] if t <= 0 else C["white"])
            c = ws.cell(row=ri, column=ci, value=val)
            _ap(c, *_plain(bg=cell_bg, bold=(ci in (1,10))))

    _scw(ws, {"A":12,"B":11,"C":13,"D":13,"E":11,"F":12,"G":12,"H":10,"I":9,"J":9})


def _build_log_vs_sim(
    wb: Workbook,
    sim_result: SimulationResult,
    log=None,   # FlightLog or None
):
    ws = wb.create_sheet("Log_vs_Sim")
    ws.sheet_view.showGridLines = False
    _merge_hdr(ws, "A1:G1",
               "Simulation vs Real Flight Log Comparison")
    ws.row_dimensions[1].height = 36

    if log is None or not HAS_MPL:
        ws.cell(row=3, column=1, value=(
            "No real flight log loaded." if log is None else
            "matplotlib not available for chart generation."
        ))
        return

    import numpy as np
    fig, axes = plt.subplots(2, 2, figsize=(14, 9))
    fig.suptitle("Simulation vs Real Flight Log", fontsize=13, fontweight="bold")

    t_sim  = np.array(sim_result.time_s)
    v_sim  = np.array(sim_result.voltage_v)
    soc_sim= np.array(sim_result.soc_pct)
    i_sim  = np.array(sim_result.current_a)

    t_log  = np.array(log.time_s)
    v_log  = np.array(log.voltage_v)
    i_log  = np.array(log.current_a)
    soc_log= np.array(log.soc_pct) if log.soc_pct else None

    # Voltage
    axes[0,0].plot(t_sim, v_sim, "b-", linewidth=2, label="Simulation", alpha=0.9)
    axes[0,0].plot(t_log, v_log, "r--", linewidth=1.5, label="Real log", alpha=0.8)
    axes[0,0].set_title("Terminal Voltage"); axes[0,0].set_ylabel("V (V)")
    axes[0,0].legend(fontsize=8); axes[0,0].grid(alpha=0.3)

    # Current
    axes[0,1].plot(t_sim, i_sim, "b-", linewidth=2, label="Simulation", alpha=0.9)
    axes[0,1].plot(t_log, i_log, "r--", linewidth=1.5, label="Real log", alpha=0.8)
    axes[0,1].set_title("Discharge Current"); axes[0,1].set_ylabel("I (A)")
    axes[0,1].legend(fontsize=8); axes[0,1].grid(alpha=0.3)

    # SoC
    if soc_log is not None:
        axes[1,0].plot(t_sim, soc_sim, "b-", linewidth=2, label="Simulation")
        axes[1,0].plot(t_log, soc_log, "r--", linewidth=1.5, label="Real log")
        axes[1,0].set_title("State of Charge"); axes[1,0].set_ylabel("SoC (%)")
        axes[1,0].legend(fontsize=8); axes[1,0].grid(alpha=0.3)

    # Residuals: V_sim - V_log (resampled to sim time axis)
    if len(t_log) > 10:
        v_log_interp = np.interp(t_sim, t_log, v_log,
                                  left=v_log[0], right=v_log[-1])
        residual = v_sim - v_log_interp
        axes[1,1].plot(t_sim, residual, "purple", linewidth=1.5)
        axes[1,1].axhline(0, color="black", linewidth=0.8, linestyle="--")
        axes[1,1].fill_between(t_sim, residual, alpha=0.2, color="purple")
        axes[1,1].set_title("Voltage residual (sim − log)")
        axes[1,1].set_ylabel("ΔV (V)")
        axes[1,1].grid(alpha=0.3)

        rmse = math.sqrt(float(np.mean(residual**2)))
        axes[1,1].text(0.05, 0.95, f"RMSE = {rmse:.4f} V",
                       transform=axes[1,1].transAxes, fontsize=9,
                       verticalalignment="top",
                       bbox=dict(boxstyle="round", facecolor="wheat", alpha=0.5))

    for ax in axes.flat:
        ax.set_xlabel("Time (s)")
    plt.tight_layout()

    buf = _chart_to_image(fig)
    ws.row_dimensions[3].height = 350
    _insert_image(ws, buf, "A3", width_px=700)
    ws.column_dimensions["A"].width = 100


def _build_fitted_params(
    wb: Workbook,
    fitted=None,   # FittedBatteryParams or None
    catalog_pack: Optional[BatteryPack] = None,
):
    ws = wb.create_sheet("Fitted_Params")
    ws.sheet_view.showGridLines = False
    _merge_hdr(ws, "A1:F1", "Reverse-Engineered Battery Parameters")
    ws.row_dimensions[1].height = 36

    if fitted is None:
        ws.cell(row=3, column=1, value="No flight log analysed — fitted parameters not available.")
        return

    headers = ["Parameter", "Fitted Value", "Uncertainty (±1σ)", "R² / Confidence",
               "Catalog Value", "Deviation"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        _ap(c, *_hdr(bg=C["sub_bg"], sz=10))
    ws.row_dimensions[2].height = 28

    rows = []
    if fitted.r_internal_mohm:
        fr = fitted.r_internal_mohm
        cat = catalog_pack.internal_resistance_mohm if catalog_pack else None
        dev = (fr.value - cat) / cat * 100 if cat else None
        rows.append(["R_internal (mΩ)", fr.value, fr.uncertainty,
                     f"R²={fr.r_squared:.3f}", cat,
                     f"{dev:+.1f}%" if dev is not None else "—"])
    if fitted.actual_capacity_ah:
        fr = fitted.actual_capacity_ah
        cat = catalog_pack.pack_capacity_ah if catalog_pack else None
        dev = (fr.value - cat) / cat * 100 if cat else None
        rows.append(["Actual capacity (Ah)", round(fr.value, 4), fr.uncertainty,
                     fr.notes[:30] if fr.notes else "—", cat,
                     f"{dev:+.1f}%" if dev is not None else "—"])
    if fitted.degradation_pct:
        rows.append(["Capacity degradation (%)", round(fitted.degradation_pct, 1),
                     "—", "—", "0.0%", f"{fitted.degradation_pct:+.1f}%"])
    if fitted.peukert_k:
        fr = fitted.peukert_k
        rows.append(["Peukert k", round(fr.value, 4), fr.uncertainty,
                     fr.notes[:30] if fr.notes else "—", 1.05,
                     f"{fr.value - 1.05:+.4f}"])
    if fitted.B_ohmic_K:
        fr = fitted.B_ohmic_K
        rows.append(["B_ohmic (K)", round(fr.value, 0), fr.uncertainty,
                     f"R²={fr.r_squared:.3f}", "—", "—"])
    if fitted.B_ct_K:
        fr = fitted.B_ct_K
        rows.append(["B_ct (K)", round(fr.value, 0), fr.uncertainty,
                     fr.method, "—", "—"])
    if fitted.ocv_soc_points:
        rows.append([f"OCV curve points", len(fitted.ocv_soc_points),
                     "—", f"SoC: {fitted.ocv_soc_points[0]:.0f}–{fitted.ocv_soc_points[-1]:.0f}%",
                     "—", "—"])

    for ri, row in enumerate(rows, start=3):
        ws.row_dimensions[ri].height = 22
        bg = C["white"] if ri % 2 == 0 else C["gray_l"]
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            dev_bg = (C["red_l"]   if isinstance(val, str) and val.startswith("+") and "%" in val else
                      C["green"]   if isinstance(val, str) and val.startswith("-") and "%" in val else bg)
            _ap(c, *_plain(bg=dev_bg if ci == 6 else bg, bold=(ci == 1)))

    # Warnings
    if fitted.fit_warnings:
        warn_row = len(rows) + 5
        ws.merge_cells(f"A{warn_row}:F{warn_row}")
        c_warn = ws.cell(row=warn_row, column=1, value="Fit Warnings")
        _ap(c_warn, *_hdr(bg=C["orange"], fg=C["black"], sz=10))
        for i, w in enumerate(fitted.fit_warnings, start=warn_row + 1):
            ws.row_dimensions[i].height = 18
            ws.merge_cells(f"A{i}:F{i}")
            c = ws.cell(row=i, column=1, value=f"  {w}")
            _ap(c, *_plain(bg=C["orange"], center=False))

    _scw(ws, {"A":26,"B":14,"C":14,"D":34,"E":14,"F":12})


# ── Main report entry point ───────────────────────────────────────────────────

def generate_report(
    out_path: str | Path,
    results: list[SimulationResult],
    packs: list[BatteryPack],
    mission: MissionProfile,
    uav_name: str = "",
    ambient_temp_c: float = 25.0,
    temp_sweep_temps: Optional[list[float]] = None,
    temp_sweep_results: Optional[list[SimulationResult]] = None,
    flight_log=None,
    fitted_params=None,
    primary_pack: Optional[BatteryPack] = None,
) -> Path:
    """
    Generate a complete formatted Excel report.

    Args:
        out_path             : Output file path (.xlsx)
        results              : List of SimulationResult (one per pack)
        packs                : Corresponding BatteryPack list
        mission              : MissionProfile used in simulation
        uav_name             : UAV configuration name
        ambient_temp_c       : Reference ambient temperature
        temp_sweep_temps     : Optional list of temperatures for sweep sheet
        temp_sweep_results   : Corresponding SimulationResult list
        flight_log           : Optional FlightLog for Log_vs_Sim sheet
        fitted_params        : Optional FittedBatteryParams for Fitted_Params sheet
        primary_pack         : Primary pack (used for detailed sheets; defaults to results[0])

    Returns:
        Path to the generated report
    """
    out_path = Path(out_path)
    wb = Workbook()

    primary_result = results[0] if results else None
    primary_pack   = primary_pack or (packs[0] if packs else None)

    meta = {
        "packs":     ", ".join(p.battery_id for p in packs),
        "mission":   mission.mission_id,
        "uav":       uav_name,
        "temp_c":    ambient_temp_c,
    }
    _build_cover(wb, meta)

    if primary_result and primary_pack:
        _build_mission_summary(wb, primary_result, primary_pack, mission)

    if results and packs:
        _build_battery_scorecard(wb, results, packs,
                                  mission.mission_id, ambient_temp_c)

    if temp_sweep_temps and temp_sweep_results and primary_pack:
        _build_temp_sensitivity(wb, temp_sweep_temps, temp_sweep_results,
                                 primary_pack.battery_id, mission.mission_id)

    if primary_result:
        _build_log_vs_sim(wb, primary_result, flight_log)

    _build_fitted_params(wb, fitted_params, primary_pack)

    wb.save(out_path)
    print(f"✓ Report saved → {out_path}")
    return out_path
