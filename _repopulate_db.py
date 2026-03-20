"""
_repopulate_db.py
Delete all TATTU_ and GREPOW_ entries from battery_db.xlsx,
then re-scrape both brands and save clean results.

Run from project root:
    python _repopulate_db.py
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
sys.stdout.reconfigure(encoding='utf-8')

from scripts.scrape_batteries import (
    delete_from_db, save_to_db, scrape_tattu, scrape_grepow
)

DB = 'battery_db.xlsx'

# ── 1. Find all existing scraped IDs ─────────────────────────────────────────
try:
    from openpyxl import load_workbook
    wb = load_workbook(DB)
    sheet_name = next((n for n in wb.sheetnames
                       if 'battery' in n.lower() and 'catalog' in n.lower()), None)
    ws = wb[sheet_name]
    to_delete = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        val = row[0]
        if val and str(val).startswith(('TATTU_', 'GREPOW_')):
            to_delete.append(str(val).strip())
    print(f'Found {len(to_delete)} existing scraped entries to delete')
except Exception as e:
    print(f'ERROR reading DB: {e}')
    sys.exit(1)

# ── 2. Delete them ────────────────────────────────────────────────────────────
if to_delete:
    deleted = delete_from_db(to_delete, db_path=DB)
    print(f'Deleted {deleted} rows\n')
else:
    print('Nothing to delete\n')

# ── 3. Scrape fresh data ──────────────────────────────────────────────────────
print('Starting Tattu scrape...')
tattu_rows = scrape_tattu(verbose=True)

print('Starting Grepow scrape...')
grepow_rows = scrape_grepow(verbose=True)

all_rows = tattu_rows + grepow_rows
print(f'\nTotal scraped: {len(all_rows)} packs '
      f'({len(tattu_rows)} Tattu, {len(grepow_rows)} Grepow)')

# ── 4. Save ───────────────────────────────────────────────────────────────────
if all_rows:
    written = save_to_db(all_rows, db_path=DB, skip_existing=False, verbose=True)
    print(f'\nDone — {written} packs written to {DB}')
else:
    print('\nNo packs scraped — DB unchanged')
